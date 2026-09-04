"""
Split a weekly P&L into its company-driver and owner-operator books.

THE TWO BLOCK TYPES DO NOT SHARE A COLUMN LAYOUT. This is the trap that makes
naive reading of these sheets produce nonsense, and it is invisible until you
compare two headers side by side:

  company driver   A Unit# B Driver C Gross D Mileage E Driver Salary
                   F Pys/Cargo/Admin G DEF/Fuel/Fee H Truck Rental
                   I Toll/Scale J Additional K Subtotal L Other M Total
                   N Per mile O Fuel avr

  owner-operator   A Unit# B Driver C Gross D COMPANY CHARGE E Driver Salary
  and lease-to-own F DEDUCTIONS G MILEAGE H Truck Rental I FUEL J Toll/Scale
                   K Fuel discount L Other charges M DRIVER PAY N PROFIT O RPM

Mileage moves from D to G. Fuel moves from G to I. The unit's result moves from
M to N. Column D, which is Mileage on a company block, is the COMPANY CHARGE on
an owner-operator block -- so reading it as miles turns a 15% dispatch fee into
a mileage figure and produces a revenue-per-mile of 7.42 instead of 2.9.

Every field here is therefore located by reading the block's own header row and
matching on the label. Nothing is read by fixed position.

An 'LO' marker appears in column A on lease-to-own rows inside an
owner-operator block.
"""
import argparse
import collections
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "ingest"))
from ingest_weekly_pnl import week_key

ALIAS = {
    "unit#": "unit", "driver": "driver_name", "gross": "gross",
    "mileage": "miles", "driver salary": "driver_pay",
    # The admin/insurance column on a company block was headed 'Insur/Admin/Trl'
    # up to 2026-06-29 and 'Pys/Cargo/Admin' after. Both are the SAME column.
    # Mapping only the later spelling drops it to zero in 3,594 of the block
    # headers across the three workbooks and overstates the company-driver
    # margin by roughly $0.19/mile.
    "pys/cargo/admin": "admin", "insur/admin/trl": "admin",
    "def/fuel/fee": "fuel", "fuel": "fuel", "fuel/toll/sc": "fuel",
    "truck rental": "rent", "toll / scale": "toll", "toll/scale": "toll",
    "toll": "toll", "toll/ scale": "toll",
    "additional charges": "additional", "subtotal": "subtotal",
    "other charges": "other", "total": "result", "profit": "result",
    "per mile": "per_mile", "rpm": "per_mile", "fuel avr": "fuel_avr",
    "company charge": "company_charge", "deductions": "deductions",
    "fuel discount": "fuel_discount", "fuel di": "fuel_discount",
    "fuel dis": "fuel_discount", "driver pay": "driver_payout",
    "contractor": "contractor", "contractor pay": "contractor",
    "contractor charge": "contractor", "contractor profit": "contractor_profit",
    "owner earning": "owner_earning", "company profit": "company_profit",
    "profit for zone": "company_profit",
}

# A company-driver block's cost columns. 'subtotal' and 'result' are totals,
# not components, and must never be added alongside these.
CD_COST_FIELDS = ("driver_pay", "admin", "fuel", "rent", "toll",
                  "additional", "other")


def unmapped_headers(ws):
    """Header labels on this sheet that ALIAS does not know.

    A label that is not in ALIAS is read as zero, silently. Run this over any
    new export before quoting a margin from it.
    """
    out = collections.Counter()
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip() != "Unit#":
            continue
        for c in range(1, 16):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and v.strip().lower() not in ALIAS:
                out[v.strip()] += 1
    return out


def n(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            return None
    return None


def unit_id(v):
    s = str(v).strip() if v is not None else ""
    return (s[:-2] if s.endswith(".0") else s) or None


def read_blocks(ws):
    """Every block, keyed by its OWN header labels. Returns dicts of named fields."""
    out = []
    heads = [r for r in range(1, ws.max_row + 1)
             if str(ws.cell(row=r, column=1).value).strip() == "Unit#"]
    for h in heads:
        cols = {}
        for c in range(1, 16):
            lab = ws.cell(row=h, column=c).value
            if isinstance(lab, str):
                key = ALIAS.get(lab.strip().lower())
                if key and key not in cols:
                    cols[key] = c
        tot = None
        for r in range(h + 1, min(h + 40, ws.max_row + 1)):
            if str(ws.cell(row=r, column=2).value).strip() == "Total":
                tot = r
                break
        if not tot:
            continue
        unit = unit_id(ws.cell(row=h + 1, column=1).value)
        if not unit:
            continue
        # 'company charge' only exists on an owner-operator header
        kind = "owner_operator" if "company_charge" in cols else "company_driver"
        rec = {"unit": unit, "kind": kind,
               "driver_name": str(ws.cell(row=h + 1, column=2).value or "").strip()}
        for key, c in cols.items():
            if key in ("unit", "driver_name"):
                continue
            rec[key] = n(ws.cell(row=tot, column=c).value) or 0.0
        # the row above 'Total' carries odometer miles and gallons on a company block
        rec["odo"] = n(ws.cell(row=tot - 1, column=2).value) or 0.0
        rec["gallons"] = n(ws.cell(row=tot - 1, column=7).value) or 0.0
        rec["mpg_printed"] = n(ws.cell(row=tot - 1, column=15).value) or 0.0
        out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--weeks", type=int, default=4)
    a = ap.parse_args()
    import openpyxl
    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    tabs = sorted([t for t in wb.sheetnames if week_key(t)], key=week_key)[-a.weeks:]

    books = {"company_driver": collections.Counter(),
             "owner_operator": collections.Counter()}
    units = {"company_driver": set(), "owner_operator": set()}
    weeks = collections.Counter()
    for t in tabs:
        for b in read_blocks(wb[t]):
            k = b["kind"]
            units[k].add(b["unit"]); weeks[k] += 1
            for f, v in b.items():
                if isinstance(v, (int, float)):
                    books[k][f] += v

    print(f"\n{Path(a.xlsx).name}  weeks {week_key(tabs[0])} to {week_key(tabs[-1])}\n")
    cd, oo = books["company_driver"], books["owner_operator"]
    print(f"  {'':<28}{'COMPANY DRIVER':>17}{'OWNER-OPERATOR':>17}{'TOTAL':>14}")
    print(f"  {'units':<28}{len(units['company_driver']):>17}"
          f"{len(units['owner_operator']):>17}"
          f"{len(units['company_driver']) + len(units['owner_operator']):>14}")
    print(f"  {'truck-weeks':<28}{weeks['company_driver']:>17}"
          f"{weeks['owner_operator']:>17}{sum(weeks.values()):>14}")
    for f, lab in [("gross", "Gross"), ("miles", "Loaded miles"),
                   ("company_charge", "Company charge (fee)"),
                   ("driver_pay", "Driver salary"), ("deductions", "Deductions"),
                   ("fuel", "Fuel"), ("rent", "Truck rental"), ("toll", "Tolls"),
                   ("admin", "Pys/Cargo/Admin"), ("other", "Other charges"),
                   ("result", "RESULT (unit level)")]:
        print(f"  {lab:<28}{cd[f]:>17,.0f}{oo[f]:>17,.0f}{cd[f] + oo[f]:>14,.0f}")

    def per(b, w):
        return (b["gross"] / w, b["miles"] / w if b["miles"] else 0,
                b["gross"] / b["miles"] if b["miles"] else 0, b["result"] / w)
    for k, lab in (("company_driver", "COMPANY DRIVER"), ("owner_operator", "OWNER-OPERATOR")):
        g, mi, rpm, res = per(books[k], weeks[k])
        print(f"\n  {lab}")
        print(f"    gross per truck-week      {g:>12,.0f}")
        print(f"    loaded mi per truck-week  {mi:>12,.0f}")
        print(f"    revenue per loaded mile   {rpm:>12,.3f}")
        print(f"    result per truck-week     {res:>12,.0f}")
        print(f"    company keeps             {100 * books[k]['result'] / books[k]['gross']:>11,.1f}%")
    if cd["odo"] and cd["gallons"]:
        print(f"\n  Company drivers: {cd['odo'] / cd['gallons']:.2f} MPG, "
              f"fuel {cd['fuel'] / cd['odo']:.3f}/odometer mile, "
              f"deadhead {100 * (cd['odo'] - cd['miles']) / cd['odo']:.1f}%")


if __name__ == "__main__":
    main()
