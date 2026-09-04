"""Why a company's weekly profit moves: a bridge from the best weeks to the last.

The workbooks answer "how much" on every tab. They do not answer "why", because
the bottom line moves for four unrelated reasons at once -- fleet size, miles
per truck, margin per mile, and overhead -- and a single net-profit series hides
all four inside one number. This module separates them.

THE DECOMPOSITION

    net  =  CD result  +  OO result  -  overhead  +  unallocated

    CD result  =  trucks  x  loaded miles per truck  x  result per loaded mile

The second line is the one that matters: a company-driver truck earns nothing
by existing, and almost every cost it carries -- rent, insurance, a driver on
weekly pay -- is charged whether it rolls or not. So miles per truck and result
per mile have to be read together. A fleet that keeps its rate per mile while
losing miles per truck is losing money at an unchanged rate.

'unallocated' is the gap between the sheet's own 'Net profit' and what the unit
blocks plus the overhead row account for. It is NOT zero (it ran $0.6k-$21k a
week over the 27 weeks to 2026-08-24) and it is not insurance, which is carried
separately and tested. It is reported, never buried in another line.

FOUR TRAPS, each of which silently changes the answer

1. LABEL DRIFT IN THE BLOCK HEADERS. The admin/insurance column on a company
   block is headed 'Insur/Admin/Trl' through 2026-06-29 and 'Pys/Cargo/Admin'
   after it. They are the same column. Knowing only the later spelling reads the
   earlier weeks as zero and invents a new cost appearing in July -- it flatters
   the early margin by about $0.13/mile and manufactures a $464/truck 'new
   charge'. 'Toll', 'Toll/ Scale' and 'Toll / Scale' are likewise one column,
   and 'Fuel di'/'Fuel dis' are 'Fuel discount'. unmapped_headers() exists so a
   new spelling fails loudly instead of reading as zero.

2. TWO BLOCK LAYOUTS. Company-driver and owner-operator blocks do not share a
   column order -- see analysis/xtrack_diagnosis.py. Everything is read by the
   block's own header.

3. THE ITEMISATION OF 'Other expenses total' IS ON ITS OWN ROWS, label and
   value side by side, and the block moves column between weeks. It is located
   by testing candidate columns until the items sum to the stated total. In 26
   of 27 XTRACK weeks the tie is exact; the exception is 2026-08-17, where
   'Freight Expenses' ($59.60) sits outside the stated total.

4. TWO MILEAGE DEFINITIONS. 'Total mileage' is loaded miles for every truck;
   'Total Odometer mileage' is all miles for company drivers only. Revenue per
   mile takes loaded miles; fuel economy takes odometer miles. Deadhead is a
   company truck's own odometer against its own loaded miles -- both come off
   the same block, so it is measurable per truck and per week.
"""
import argparse
import collections
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent / "ingest"))
sys.path.insert(0, str(Path(__file__).parent))
from ingest_weekly_pnl import week_key, num, labeled, overhead_row, WANT
from xtrack_diagnosis import read_blocks, unmapped_headers, CD_COST_FIELDS

OVERHEAD_KEYS = ("us_office", "owners", "tashkent", "other")


def other_itemisation(ws, total):
    """Split 'Other expenses total' into its named lines.

    Returns (items, residual). The label sits in one column and its amount in
    the column immediately to its left, on the SAME row -- but which pair of
    columns is not fixed, so candidates are tried until the items reproduce the
    stated total. residual is stated_total - sum(items): non-zero means a line
    on the sheet sits outside the total the sheet itself prints.
    """
    best = None
    for lc in range(18, 25):
        items = {}
        for r in range(1, 50):
            lab = ws.cell(row=r, column=lc).value
            if not isinstance(lab, str) or not lab.strip():
                continue
            v = num(ws.cell(row=r, column=lc - 1).value)
            if v is not None:
                k = lab.strip().rstrip()
                items[k] = items.get(k, 0.0) + v
        if not items:
            continue
        gap = (total or 0.0) - sum(items.values())
        if best is None or abs(gap) < abs(best[1]):
            best = (items, gap)
        if abs(gap) < 1.0:
            return items, gap
    return best if best else ({}, total or 0.0)


def week_record(ws):
    """One week, from the unit blocks up to the sheet's own bottom line."""
    panel = labeled(ws, WANT)
    oh = overhead_row(ws)
    blocks = read_blocks(ws)
    cd = [b for b in blocks if b["kind"] == "company_driver"]
    oo = [b for b in blocks if b["kind"] == "owner_operator"]
    s = lambda rows, f: sum(r.get(f, 0.0) or 0.0 for r in rows)
    items, item_gap = other_itemisation(ws, panel.get("other_exp"))
    pick = lambda pred: sum(v for k, v in items.items() if pred(k))
    rec = {
        "net": num(oh.get("net_profit")) or 0.0,
        "gross": panel.get("gross") or 0.0,
        "insurance": panel.get("insurance") or 0.0,
        "overhead": sum(abs(oh.get(k) or 0.0) for k in OVERHEAD_KEYS),
        "tashkent": abs(oh.get("tashkent") or 0.0),
        "other_charges": abs(oh.get("other") or 0.0),
        "cd_trucks": len(cd),
        "cd_idle": sum(1 for b in cd if (b.get("gross") or 0) <= 0),
        "oo_trucks": len(oo),
        "cd_gross": s(cd, "gross"), "cd_result": s(cd, "result"),
        "cd_miles": s(cd, "miles"), "cd_odo": s(cd, "odo"),
        "cd_gallons": s(cd, "gallons"),
        "oo_gross": s(oo, "gross"), "oo_result": s(oo, "result"),
        "maintenance": pick(lambda k: "aintenance" in k),
        "trailer": pick(lambda k: "railer" in k or "Reefer" in k),
        "factoring": pick(lambda k: "Factoring" in k),
        "stl": pick(lambda k: "STL" in k),
        "item_gap": item_gap,
        "unit_gross": sum((b.get("gross") or 0.0) for b in blocks),
        "unmapped": dict(unmapped_headers(ws)),
    }
    for f in CD_COST_FIELDS:
        rec["cd_" + f] = s(cd, f)
    rec["unallocated"] = rec["net"] - (rec["cd_result"] + rec["oo_result"] - rec["overhead"])
    return rec


def load(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    tabs = sorted([t for t in wb.sheetnames if week_key(t)], key=week_key)
    return {week_key(t): week_record(wb[t]) for t in tabs}


def controls(weeks, tol=1.0):
    """Every check that must hold before a number here is quoted anywhere."""
    fails = []
    for wk, d in sorted(weeks.items()):
        if d["unmapped"]:
            fails.append((wk, "unmapped block headers", d["unmapped"]))
        if abs(d["unit_gross"] - d["gross"]) > tol:
            fails.append((wk, "unit blocks vs panel gross",
                          round(d["unit_gross"] - d["gross"], 2)))
        costs = sum(d["cd_" + f] for f in CD_COST_FIELDS)
        if abs(d["cd_gross"] - costs - d["cd_result"]) > tol:
            fails.append((wk, "CD gross - costs vs CD result",
                          round(d["cd_gross"] - costs - d["cd_result"], 2)))
        if abs(d["item_gap"]) > tol:
            fails.append((wk, "other-expense itemisation vs stated total",
                          round(d["item_gap"], 2)))
    return fails


def period(weeks, keys):
    """Weekly averages over a set of weeks, plus the ratios that carry meaning."""
    n = len(keys)
    a = lambda f: sum(weeks[k][f] for k in keys) / n
    p = {f: a(f) for f in weeks[keys[0]] if f not in ("unmapped",)}
    p["weeks"] = n
    p["cd_miles_per_truck"] = p["cd_miles"] / p["cd_trucks"] if p["cd_trucks"] else 0.0
    p["cd_rpm"] = p["cd_gross"] / p["cd_miles"] if p["cd_miles"] else 0.0
    p["cd_result_per_mile"] = p["cd_result"] / p["cd_miles"] if p["cd_miles"] else 0.0
    p["cd_cost_per_mile"] = sum(p["cd_" + f] for f in CD_COST_FIELDS) / p["cd_miles"]
    p["cd_deadhead"] = (p["cd_odo"] - p["cd_miles"]) / p["cd_odo"] if p["cd_odo"] else 0.0
    p["cd_mpg"] = p["cd_odo"] / p["cd_gallons"] if p["cd_gallons"] else 0.0
    p["cd_dollars_per_gallon"] = p["cd_fuel"] / p["cd_gallons"] if p["cd_gallons"] else 0.0
    p["oo_result_per_truck"] = p["oo_result"] / p["oo_trucks"] if p["oo_trucks"] else 0.0
    return p


def bridge(a, b):
    """Net profit, period a -> period b, one line per cause. Adds up exactly."""
    return [
        ("company-driver result", b["cd_result"] - a["cd_result"]),
        ("owner-operator result", b["oo_result"] - a["oo_result"]),
        ("overhead", -(b["overhead"] - a["overhead"])),
        ("unallocated", b["unallocated"] - a["unallocated"]),
    ]


def cd_factors(a, b):
    """CD result = trucks x miles/truck x result/mile, one factor at a time."""
    t0, m0, r0 = a["cd_trucks"], a["cd_miles_per_truck"], a["cd_result_per_mile"]
    t1, m1, r1 = b["cd_trucks"], b["cd_miles_per_truck"], b["cd_result_per_mile"]
    s0 = t0 * m0 * r0
    s1 = t1 * m0 * r0
    s2 = t1 * m1 * r0
    s3 = t1 * m1 * r1
    return [("fewer trucks (%.1f -> %.1f)" % (t0, t1), s1 - s0),
            ("miles per truck (%,.0f -> %,.0f)".replace(",", "") % (m0, m1), s2 - s1),
            ("result per loaded mile ($%.3f -> $%.3f)" % (r0, r1), s3 - s2)]


def fuel_walk(a, b):
    """Fuel per LOADED mile = ($/gal) / mpg x (odometer / loaded). Price, economy, deadhead."""
    g0, e0 = a["cd_dollars_per_gallon"], a["cd_mpg"]
    g1, e1 = b["cd_dollars_per_gallon"], b["cd_mpg"]
    d0 = a["cd_odo"] / a["cd_miles"]
    d1 = b["cd_odo"] / b["cd_miles"]
    base = g0 / e0 * d0
    p = g1 / e0 * d0
    m = g1 / e1 * d0
    f = g1 / e1 * d1
    return [("fuel price ($%.3f -> $%.3f /gal)" % (g0, g1), p - base),
            ("fuel economy (%.2f -> %.2f mpg)" % (e0, e1), m - p),
            ("deadhead (%.1f%% -> %.1f%%)" % (100 * a["cd_deadhead"], 100 * b["cd_deadhead"]), f - m)]


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--span", type=int, default=6,
                    help="weeks in the 'best' and 'last' comparison windows")
    ap.add_argument("--json", help="write the weekly series here")
    a = ap.parse_args()

    weeks = load(a.xlsx)
    ks = sorted(weeks)
    fails = controls(weeks)
    print(f"{len(ks)} weeks, {ks[0]} .. {ks[-1]}")
    if fails:
        print(f"CONTROLS FAILED ({len(fails)}):")
        for wk, what, detail in fails:
            print(f"  {wk}  {what}: {detail}")
    else:
        print("controls: all pass")
    print()

    hdr = ("week", "net", "gross", "CD", "idle", "mi/tk", "RPM", "$/mi",
           "CDres", "OOres", "maint", "other", "tash")
    print("".join(f"{h:>{w}}" for h, w in zip(hdr, (12, 9, 10, 4, 5, 7, 6, 6, 9, 9, 8, 9, 8))))
    for k in ks:
        d = weeks[k]
        mi = d["cd_miles"] / d["cd_trucks"] if d["cd_trucks"] else 0
        rpm = d["cd_gross"] / d["cd_miles"] if d["cd_miles"] else 0
        cpm = sum(d["cd_" + f] for f in CD_COST_FIELDS) / d["cd_miles"] if d["cd_miles"] else 0
        print(f"{k:>12}{d['net']:>9,.0f}{d['gross']:>10,.0f}{d['cd_trucks']:>4}"
              f"{d['cd_idle']:>5}{mi:>7,.0f}{rpm:>6.2f}{cpm:>6.2f}"
              f"{d['cd_result']:>9,.0f}{d['oo_result']:>9,.0f}{d['maintenance']:>8,.0f}"
              f"{d['other_charges']:>9,.0f}{d['tashkent']:>8,.0f}")

    best = sorted(ks, key=lambda k: weeks[k]["net"], reverse=True)[:a.span]
    A = period(weeks, sorted(best))
    B = period(weeks, ks[-a.span:])
    print(f"\nbest {a.span} weeks by net: {', '.join(sorted(best))}")
    print(f"last {a.span} weeks:         {', '.join(ks[-a.span:])}")

    def walk(title, rows, start, end):
        print(f"\n{title}")
        print(f"  {'from':<44}{start:>12,.0f}")
        for lab, v in rows:
            print(f"  {lab:<44}{v:>12,.0f}")
        print(f"  {'to':<44}{end:>12,.0f}   (check {start + sum(v for _, v in rows) - end:+,.0f})")

    walk("NET PROFIT PER WEEK", bridge(A, B), A["net"], B["net"])
    walk("COMPANY-DRIVER RESULT PER WEEK", cd_factors(A, B), A["cd_result"], B["cd_result"])

    print("\nCOMPANY-DRIVER COST PER LOADED MILE")
    print(f"  {'revenue per loaded mile':<44}{A['cd_rpm']:>10.3f}{B['cd_rpm']:>10.3f}"
          f"{B['cd_rpm'] - A['cd_rpm']:>+10.3f}")
    for f in CD_COST_FIELDS:
        x, y = A["cd_" + f] / A["cd_miles"], B["cd_" + f] / B["cd_miles"]
        print(f"  {f:<44}{x:>10.3f}{y:>10.3f}{y - x:>+10.3f}")
    print(f"  {'= result per loaded mile':<44}{A['cd_result_per_mile']:>10.3f}"
          f"{B['cd_result_per_mile']:>10.3f}"
          f"{B['cd_result_per_mile'] - A['cd_result_per_mile']:>+10.3f}")

    print("\nOF WHICH, FUEL PER LOADED MILE")
    for lab, v in fuel_walk(A, B):
        print(f"  {lab:<44}{v:>+10.3f}")

    print("\nOVERHEAD PER WEEK")
    for f in ("tashkent", "other_charges", "overhead", "insurance", "maintenance", "trailer"):
        print(f"  {f:<44}{A[f]:>10,.0f}{B[f]:>10,.0f}{B[f] - A[f]:>+10,.0f}")

    if a.json:
        json.dump(weeks, open(a.json, "w"), indent=1)


if __name__ == "__main__":
    main()
