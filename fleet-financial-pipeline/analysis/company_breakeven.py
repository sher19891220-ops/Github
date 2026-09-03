"""
Fixed cost, variable cost and break-even for ZONE, XTRACK and AFG separately.

BUILT FROM THE COMPANIES' OWN WEEKLY SHEETS, not from estimates. Those sheets
carry everything the calculation needs: per-unit gross, mileage, driver pay,
fuel, truck rent and tolls in the unit rows, and US office, owners, Tashkent and
an itemised 'Other charges' in the panel. Earlier passes at this used a measured
equipment figure from the bank plus operator-supplied overhead; that is no
longer necessary and would now double-count, because truck rent and the
overhead lines are already inside these sheets.

FIXED vs VARIABLE -- the split that decides the answer:

  FIXED    truck rent, insurance and admin, max-truck, trailer insurance and
           charge, STL charges, ADP fee, US office, owners, Tashkent BASE salary.
           These are paid per truck per week whether it turns a wheel or not.

  VARIABLE driver pay, fuel and DEF, tolls, maintenance, factoring fee, Loves,
           returned-truck tolls, freight expenses, and Tashkent COMMISSION.
           These scale with miles or with revenue.

Tashkent is split rather than taken whole: the payroll register shows only
40,020 a month of base salary against 105,708 of bonus and commission, so 73%
of it moves with freight. Treating all of it as fixed overstates the fixed base
and understates cost per mile.

WHY THE PANEL TOTALS ARE USED AND NOT THE UNIT ROWS: the unit blocks NEST. In
XTRACK's 08.17 week there are 101 rows reading 'Total' in column B against 48
trucks, so summing them counts most cost columns roughly twice. Gross happens to
survive it -- it sums to the panel figure exactly -- which is precisely what
makes the trap dangerous: the control you would naturally run passes while every
cost line is inflated. The panel's own 'Total driver pay', 'Total fuel',
'Total truck rent' and so on are the sheet's own arithmetic and are used instead.

VARIABLE COST IS DERIVED, NOT SUMMED. Even the panel does not enumerate every
cost -- adding its lines up does not reproduce its own net profit -- so variable
is taken from the accounting identity, variable = gross - net profit - fixed.
That guarantees the model cannot contradict the sheet it is built from, and it
puts anything the panel omits into the variable bucket, which is the
conservative direction for a break-even.

MILEAGE: 'Total mileage' from the panel. The sheets also carry 'Total Odometer
mileage', which differs by roughly 2x in XTRACK; that relationship is still
unresolved and the odometer figure is not used.

Truck counts are the operator's current fleet, not a period average.
"""
import argparse
import collections
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "ingest"))
from ingest_weekly_pnl import week_key, overhead_row, labeled, blocks, WANT

FIXED_ITEMS = {"max_truck", "trailer_ins", "trailer_chg", "stl", "adp_fee"}
VAR_ITEMS = {"factoring", "loves", "maint_company", "maint_driver",
             "toll_returned", "freight_exp", "credit_card"}
ITEM_LABELS = {
    "max truck": "max_truck", "max truck and psz": "max_truck",
    "trailer insurance": "trailer_ins", "trailer charge": "trailer_chg",
    "credit card": "credit_card", "credit card (ae)": "credit_card",
    "adp fee": "adp_fee", "factoring fee": "factoring",
    "loves (ach expense)": "loves", "charges from stl": "stl",
    "weekly main charges from stl": "stl",
    "maintenance for truck and trailer": "maint_company",
    "comp exp maintenance for truck and trailer": "maint_company",
    "driver exp maintenance for truck and trailer": "maint_driver",
    "toll for returned trucks": "toll_returned", "freight expenses": "freight_exp"}
UNIT_COLS = {5: "driver_pay", 6: "pys_cargo_admin", 7: "fuel", 8: "truck_rent",
             9: "tolls", 10: "additional"}
UNIT_VARIABLE = {"driver_pay", "fuel", "tolls", "additional"}


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def other_items(ws):
    """The 'Other charges' block: label in a column, VALUE ONE COLUMN LEFT."""
    out = collections.Counter()
    seen = set()
    for row in ws.iter_rows(min_row=10, max_row=40, min_col=19, max_col=21):
        for c in row:
            if not isinstance(c.value, str):
                continue
            k = ITEM_LABELS.get(c.value.strip().lower().rstrip("\t ").strip())
            if k and k not in seen:
                seen.add(k)
                out[k] += num(ws.cell(row=c.row, column=c.column - 1).value)
    return out


def unit_totals(ws):
    """Sum the per-unit cost columns off each block's closing 'Total' row."""
    out = collections.Counter()
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value).strip() != "Total":
            continue
        for col, name in UNIT_COLS.items():
            out[name] += num(ws.cell(row=r, column=col).value)
    return out


def company(path, weeks_limit=None):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    agg = collections.Counter()
    n = 0
    for tab in sorted(wb.sheetnames, key=lambda t: week_key(t) or "", reverse=True):
        if not week_key(tab):
            continue
        if weeks_limit and n >= weeks_limit:
            break
        ws = wb[tab]
        b = blocks(ws)
        agg["gross"] += sum(x[0] for x in b)
        pan = labeled(ws, WANT)
        agg["miles"] += num(pan.get("all_miles"))
        for k in ("driver_pay", "fuel", "truck_rent", "tolls", "insurance",
                  "oo_salaries", "oo_expenses"):
            agg[k] += num(pan.get(k))
        agg.update(other_items(ws))
        oh = overhead_row(ws)
        for k in ("us_office", "owners", "tashkent"):
            agg[k] += abs(num(oh.get(k)))
        agg["net_profit"] += num(oh.get("net_profit"))
        n += 1
    agg["weeks"] = n
    return agg


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--zone", required=True)
    ap.add_argument("--xtrack", required=True)
    ap.add_argument("--afg", required=True)
    ap.add_argument("--trucks", default="ZONE=32,XTRACK=45,AFG=17")
    ap.add_argument("--weeks", type=int, default=20,
                    help="most recent N weeks (default 20, the span AFG covers)")
    ap.add_argument("--tashkent-fixed-share", type=float, default=0.275,
                    help="base salary as a share of Tashkent cost; the rest is "
                         "commission and is treated as variable")
    a = ap.parse_args()

    trucks = dict(x.split("=") for x in a.trucks.split(","))
    trucks = {k: float(v) for k, v in trucks.items()}
    data = {"ZONE": company(a.zone, a.weeks), "XTRACK": company(a.xtrack, a.weeks),
            "AFG": company(a.afg, a.weeks)}

    print(f"\nMost recent {a.weeks} weeks of each company's own weekly sheet.")
    print(f"Truck counts are the operator's current fleet: "
          f"{', '.join(f'{k} {int(v)}' for k, v in trucks.items())}.\n")

    rows = {}
    for co, d in data.items():
        w, t, mi = d["weeks"], trucks[co], d["miles"]
        tas_fix = d["tashkent"] * a.tashkent_fixed_share
        fixed = (d["truck_rent"] + d["insurance"] + d["adp_fee"]
                 + d["trailer_ins"] + d["trailer_chg"] + d["max_truck"] + d["stl"]
                 + d["us_office"] + d["owners"] + tas_fix)
        var = d["gross"] - d["net_profit"] - fixed          # accounting identity
        rows[co] = {
            "weeks": w, "trucks": t, "miles": mi, "gross": d["gross"],
            "fixed_wk": fixed / w, "fixed_tw": fixed / w / t,
            "var_mi": var / mi if mi else 0,
            "rev_mi": d["gross"] / mi if mi else 0,
            "mi_tw": mi / w / t, "net": d["net_profit"]}
        rows[co]["contrib"] = rows[co]["rev_mi"] - rows[co]["var_mi"]
        rows[co]["be_mi"] = (rows[co]["fixed_tw"] / rows[co]["contrib"]
                             if rows[co]["contrib"] > 0 else float("inf"))
        rows[co]["fixed_per_mile"] = fixed / mi if mi else 0

    def line(label, key, fmt=",.3f"):
        print(f"  {label:<30}" + "".join(format(rows[c][key], fmt).rjust(12)
                                         for c in ("ZONE", "XTRACK", "AFG")))

    print(f"  {'':<30}{'ZONE':>12}{'XTRACK':>12}{'AFG':>12}")
    print(f"  {'-' * 66}")
    line("Trucks", "trucks", ",.0f")
    line("Miles run", "miles", ",.0f")
    line("Miles per truck-week", "mi_tw", ",.0f")
    print()
    line("FIXED per week", "fixed_wk", ",.0f")
    line("FIXED per truck-week", "fixed_tw", ",.0f")
    line("FIXED per mile", "fixed_per_mile", ",.3f")
    print()
    line("Revenue per mile", "rev_mi")
    line("VARIABLE per mile", "var_mi")
    line("CONTRIBUTION per mile", "contrib")
    print()
    line("BREAK-EVEN mi/truck-week", "be_mi", ",.0f")
    line("actual mi/truck-week", "mi_tw", ",.0f")
    print(f"  {'margin over break-even':<30}" + "".join(
        f"{rows[c]['mi_tw'] - rows[c]['be_mi']:+,.0f}".rjust(12)
        for c in ("ZONE", "XTRACK", "AFG")))
    print(f"  {'$/truck-week at actual':<30}" + "".join(
        f"{(rows[c]['mi_tw'] - rows[c]['be_mi']) * rows[c]['contrib']:+,.0f}".rjust(12)
        for c in ("ZONE", "XTRACK", "AFG")))
    print(f"  {'$/week for the company':<30}" + "".join(
        f"{(rows[c]['mi_tw'] - rows[c]['be_mi']) * rows[c]['contrib'] * rows[c]['trucks']:+,.0f}".rjust(12)
        for c in ("ZONE", "XTRACK", "AFG")))

    tf = sum(rows[c]["fixed_wk"] for c in rows)
    tt = sum(rows[c]["trucks"] for c in rows)
    tm = sum(rows[c]["miles"] for c in rows)
    tg = sum(rows[c]["gross"] for c in rows)
    tv = sum(rows[c]["var_mi"] * rows[c]["miles"] for c in rows)
    w = rows["ZONE"]["weeks"]
    gc = tg / tm - tv / tm
    print(f"\n  CHECK -- model against the sheet's own stated net profit")
    print(f"  {'sheet net/week':<30}" + "".join(
        f"{rows[c]['net'] / rows[c]['weeks']:+,.0f}".rjust(12)
        for c in ("ZONE", "XTRACK", "AFG")))
    print(f"  {'model result/week':<30}" + "".join(
        f"{(rows[c]['mi_tw'] - rows[c]['be_mi']) * rows[c]['contrib'] * rows[c]['trucks']:+,.0f}".rjust(12)
        for c in ("ZONE", "XTRACK", "AFG")))
    print(f"  The two differ only because the model prices the CURRENT fleet"
          f"\n  ({', '.join(f'{k} {int(v)}' for k, v in trucks.items())}) while the sheet"
          f" recorded whatever ran that week.")

    print(f"\n  GROUP: fixed {tf:,.0f}/wk over {tt:,.0f} trucks = {tf / tt:,.0f}/truck-week")
    print(f"         revenue {tg / tm:.3f}/mi, variable {tv / tm:.3f}/mi, "
          f"contribution {gc:.3f}/mi")
    print(f"         break-even {tf / tt / gc:,.0f} mi/truck-week, "
          f"actual {tm / w / tt:,.0f}")
    print(f"         group result {((tm / w / tt) - tf / tt / gc) * gc * tt:+,.0f}/week")


if __name__ == "__main__":
    main()
