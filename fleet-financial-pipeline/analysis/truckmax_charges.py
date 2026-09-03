"""
Truck Max shop charges, segregated by what the unit IS and who paid.

AMOUNT: 'Inv amount' is blank on 50 rows that nonetheless carry a figure in a
payer column -- 73,118.63 of real charges. So the amount is 'Inv amount' where
it is filled and the payer total otherwise, and the two are reconciled rather
than one being taken on faith.

The sheet already names the payer in three columns -- Zone, Iron Lease, Driver --
and each row carries either a Truck number or a Trailer number, never both. So
the segregation the operator wants is a matter of joining the unit number to the
fleet lists, not of guessing from the free-text 'Issue' field.

FLEET CLASSES, per the operator:
  iron_lease    trucks Iron Lease owns and leases to the operating companies
  investor      1365, 1596, 3898, 4546 -- owned by an outside investor; we find
                the driver, run the truck, take a 10% service fee and pass the
                remainder back. Not our asset and not Iron Lease's.
  lease_to_own  drivers buying the truck from us, shown as 'rented' until paid
  sold          4851 -- paid off, out of our possession, must not count as fleet
  trailer       rows with a trailer number instead of a truck number
  zone_stl      the 80xx/81xx block -- Zone bought these from STL on a
                lease-to-purchase option. Zone's own, not Iron Lease's.
  other         a truck on none of the lists: an outside lease (Ryder, Penske)
                or a unit not yet recorded

THE PAYER COLUMNS ARE A BILLING ADDRESS, NOT THE COST BEARER. Truck Max invoices
Zone for almost everything, and the fleet and accounting teams redistribute the
charges to the right company afterwards. So a figure in the 'Zone' column means
Zone was BILLED, not that Zone ultimately bore it -- and this file has no record
of the redistribution. Any per-company conclusion drawn from these columns would
be measuring the invoicing convention rather than the economics.

Amounts come from 'Inv amount'. The payer columns are checked against it: where
they disagree the row is reported rather than quietly taking one side.
"""
import argparse
import collections
from pathlib import Path

IRON_LEASE = {"15862", "15909", "15739", "4772", "15852", "8482", "7605", "6799",
              "9859", "6867", "2639", "4716", "1489", "1431", "1500", "1722",
              "1662", "1568", "1645", "4727", "3773", "5269", "4928", "5007",
              "5026", "4937", "1542", "4713", "1471", "4709", "4549", "4553"}
INVESTOR = {"1365", "1596", "3898", "4546"}
LEASE_TO_OWN = {"2703", "8671", "4864", "4857", "1564", "4718", "1682", "5413"}
SOLD = {"4851"}
# Bought by Zone from STL on a lease-to-purchase option. Zone's own trucks, not
# Iron Lease's and not an outside lease.
ZONE_STL = {"8033", "8083", "8091", "8092", "8093", "8094",
            "8130", "8131", "8132", "8133", "8136"}

# Keying errors in the source, corrected here rather than by editing the file, so
# the change stays visible and reversible. Each needs the operator's confirmation
# before it goes in.
UNIT_CORRECTIONS = {
    # INV-484, 2026-07-30, 5,761.39 paid by Iron Lease. Keyed 1543; 1543 appears
    # nowhere else in the corpus and 1542 -- a leased truck -- appears not at all.
    # The work is engine oil pan, valve adjustment, NOx sensors, DPF: mechanical,
    # truck-only, exactly Iron Lease's contractual scope. Operator confirmed.
    "1543": "1542",
}


def unit(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


def correct(u):
    """Apply a confirmed keying correction, if there is one for this unit."""
    return UNIT_CORRECTIONS.get(u, u)


def classify(truck, trailer):
    if trailer and not truck:
        return "trailer"
    if not truck:
        return "unknown"
    if truck in SOLD:
        return "sold"
    if truck in INVESTOR:
        return "investor"
    if truck in LEASE_TO_OWN:
        return "lease_to_own"
    if truck in IRON_LEASE:
        return "iron_lease"
    if truck in ZONE_STL:
        return "zone_stl"
    return "other"


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    a = ap.parse_args()
    import openpyxl
    ws = openpyxl.load_workbook(a.xlsx, data_only=True)["Sheet1"]

    hdr = [c.value for c in ws[1]]
    col = {str(h).strip(): i for i, h in enumerate(hdr, 1) if h}
    C = lambda r, n: ws.cell(row=r, column=col[n]).value if n in col else None
    def money(v):
        return float(v) if isinstance(v, (int, float)) else 0.0

    rows = []
    for r in range(2, ws.max_row + 1):
        t, tr = correct(unit(C(r, "Truck"))), unit(C(r, "Trailer"))
        amt = money(C(r, "Inv amount"))
        if not (t or tr) and not amt:
            continue
        payers = {p: money(C(r, p)) for p in ("Zone", "Iron Lease", "Driver")}
        psum = sum(payers.values())
        if amt == 0 and psum:
            amt = psum                      # invoice column left blank
        rows.append({"truck": t, "trailer": tr, "amt": amt, "labor": money(C(r, "Labor")),
                     "cls": classify(t, tr), "payers": payers,
                     "payer_sum": psum, "date": C(r, "Date")})

    print(f"\n{Path(a.xlsx).name} -- {len(rows)} charge rows")
    from datetime import datetime
    d = [r["date"] for r in rows if isinstance(r["date"], datetime)]
    if d:
        print(f"period {min(d).date()} to {max(d).date()}")
    total = sum(r["amt"] for r in rows)
    mism = [r for r in rows if abs(r["payer_sum"] - r["amt"]) > 0.01]
    print(f"total ${total:,.2f}; {len(mism)} rows still reconcile imperfectly\n")

    print("BY WHAT THE UNIT IS\n")
    print(f"  {'class':<16}{'rows':>6}{'units':>7}{'invoiced':>14}{'share':>8}{'labour':>12}")
    by = collections.defaultdict(lambda: {"n": 0, "amt": 0.0, "lab": 0.0, "u": set()})
    for r in rows:
        b = by[r["cls"]]
        b["n"] += 1; b["amt"] += r["amt"]; b["lab"] += r["labor"]
        b["u"].add(r["truck"] or r["trailer"])
    for k, v in sorted(by.items(), key=lambda x: -x[1]["amt"]):
        print(f"  {k:<16}{v['n']:>6}{len(v['u']):>7}{v['amt']:>14,.2f}"
              f"{100 * v['amt'] / total:>7.1f}%{v['lab']:>12,.2f}")
    print(f"  {'TOTAL':<16}{len(rows):>6}{'':>7}{total:>14,.2f}{100:>7.1f}%"
          f"{sum(r['labor'] for r in rows):>12,.2f}")

    print("\n\nWHO PAID, WITHIN EACH CLASS\n")
    print(f"  {'class':<16}{'Zone':>14}{'Iron Lease':>14}{'Driver':>14}{'unallocated':>14}")
    for k in sorted(by, key=lambda x: -by[x]["amt"]):
        p = collections.Counter()
        una = 0.0
        for r in rows:
            if r["cls"] != k:
                continue
            for name, v in r["payers"].items():
                p[name] += v
            if r["payer_sum"] == 0:
                una += r["amt"]
        print(f"  {k:<16}{p['Zone']:>14,.2f}{p['Iron Lease']:>14,.2f}"
              f"{p['Driver']:>14,.2f}{una:>14,.2f}")

    print("\n\n'OTHER' TRUCKS -- on none of the fleet lists\n")
    oth = collections.Counter()
    for r in rows:
        if r["cls"] == "other":
            oth[r["truck"]] += r["amt"]
    print(f"  {len(oth)} distinct units, ${sum(oth.values()):,.2f}")
    print(f"  {'unit':<10}{'invoiced':>12}")
    for u, v in oth.most_common(15):
        print(f"  {u:<10}{v:>12,.2f}")


if __name__ == "__main__":
    main()
