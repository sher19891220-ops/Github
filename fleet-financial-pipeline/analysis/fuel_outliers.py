"""
Trucks running below an MPG threshold, from the weekly P&L unit blocks.

WHERE THE TWO FUEL NUMBERS LIVE, because they are on different rows and mean
different things:

    row above each block's 'Total':  B = odometer miles, G = gallons,
                                     O = MPG  (miles / gallons)
    the 'Total' row itself:          G = fuel dollars, D = billed mileage,
                                     O = fuel COST PER MILE (dollars / odo miles)

So 'Fuel avr' is the header for BOTH columns and means MPG on one row and
dollars-per-mile on the next. Reading either row alone gives the wrong figure.

The unit number sits in column A of the block's first data row, two rows below
the 'Unit#' header.

The sheet's OWN printed MPG is what is reported, because that is the figure the
operator reads. It does not always equal miles / gallons -- unit 1471 prints
6.68 where the division gives 6.50 -- so the recomputation is run as a CHECK and
disagreements are counted and reported, not used to discard the block. Dropping
them silently would have cut the fleet from 48 units to 13.

Blocks with no fuel data at all are skipped: an owner-operator buys their own
fuel, so their block is legitimately empty. Unit numbers arrive as floats
(4937.0) and are normalised.
"""
import argparse
import collections
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "ingest"))
from ingest_weekly_pnl import week_key


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            return None
    return None


def blocks_with_fuel(ws):
    """Yield one record per unit block: unit, driver, miles, gallons, mpg, $/mi."""
    out = []
    header_rows = [r for r in range(1, ws.max_row + 1)
                   if str(ws.cell(row=r, column=1).value).strip() == "Unit#"]
    for h in header_rows:
        unit = ws.cell(row=h + 1, column=1).value
        driver = ws.cell(row=h + 1, column=2).value
        tot = None
        for r in range(h + 1, min(h + 40, ws.max_row + 1)):
            if str(ws.cell(row=r, column=2).value).strip() == "Total":
                tot = r
                break
        if not tot:
            continue
        miles = num(ws.cell(row=tot - 1, column=2).value)
        gallons = num(ws.cell(row=tot - 1, column=7).value)
        mpg = num(ws.cell(row=tot - 1, column=15).value)
        fuel = num(ws.cell(row=tot, column=7).value)
        cpm = num(ws.cell(row=tot, column=15).value)
        gross = num(ws.cell(row=tot, column=3).value)
        if not (unit and miles and gallons and mpg):
            continue
        agrees = abs(miles / gallons - mpg) < 0.05 if gallons else False
        u = str(unit).strip()
        if u.endswith(".0"):
            u = u[:-2]
        out.append({"unit": u, "driver": str(driver or "").strip(),
                    "miles": miles, "gallons": gallons, "mpg": mpg,
                    "fuel": fuel or 0.0, "cpm": cpm or 0.0, "gross": gross or 0.0,
                    "agrees": agrees})
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--weeks", type=int, default=4)
    ap.add_argument("--mpg-below", type=float, default=7.0)
    a = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    tabs = sorted([t for t in wb.sheetnames if week_key(t)], key=week_key)[-a.weeks:]

    per = collections.defaultdict(list)
    disagree = 0
    for t in tabs:
        for b in blocks_with_fuel(wb[t]):
            if not b["agrees"]:
                disagree += 1
            b["week"] = week_key(t)
            per[b["unit"]].append(b)

    print(f"\n{Path(a.xlsx).name} -- weeks {week_key(tabs[0])} to {week_key(tabs[-1])}")
    nb = sum(len(v) for v in per.values())
    print(f"{len(per)} units with fuel data across {nb} truck-weeks; on {disagree} of "
          f"them\nthe printed MPG differs from miles/gallons by more than 0.05 "
          f"({100 * disagree / nb:.0f}%).\n")

    rows = []
    for unit, bs in per.items():
        mi = sum(x["miles"] for x in bs)
        ga = sum(x["gallons"] for x in bs)
        fu = sum(x["fuel"] for x in bs)
        wmpg = (sum(x["mpg"] * x["miles"] for x in bs) / mi) if mi else 0
        rows.append({"unit": unit, "driver": bs[-1]["driver"], "weeks": len(bs),
                     "miles": mi, "gallons": ga, "mpg": wmpg,
                     "cpm": fu / mi if mi else 0, "fuel": fu,
                     "gross": sum(x["gross"] for x in bs)})
    bad = sorted([r for r in rows if r["mpg"] < a.mpg_below], key=lambda r: r["mpg"])
    good = [r for r in rows if r["mpg"] >= a.mpg_below]

    print(f"UNITS BELOW {a.mpg_below} MPG over the {a.weeks} weeks — "
          f"{len(bad)} of {len(rows)}\n")
    print(f"  {'Unit#':<8}{'Driver':<26}{'wks':>4}{'miles':>9}{'gallons':>9}"
          f"{'Fuel avr':>10}{'$/mile':>9}{'fuel $':>11}")
    for r in bad:
        print(f"  {r['unit']:<8}{r['driver'][:25]:<26}{r['weeks']:>4}{r['miles']:>9,.0f}"
              f"{r['gallons']:>9,.0f}{r['mpg']:>10,.2f}{r['cpm']:>9,.2f}{r['fuel']:>11,.0f}")

    if bad:
        fleet_mpg = sum(r["miles"] for r in rows) / sum(r["gallons"] for r in rows)
        good_mpg = (sum(r["miles"] for r in good) / sum(r["gallons"] for r in good)
                    if good else 0)
        bm = sum(r["miles"] for r in bad)
        excess_gal = bm / (sum(r['miles'] for r in bad) / sum(r['gallons'] for r in bad)) \
                     - (bm / good_mpg if good_mpg else 0)
        price = sum(r["fuel"] for r in rows) / sum(r["gallons"] for r in rows)
        print(f"\n  fleet MPG {fleet_mpg:,.2f}   units at or above threshold {good_mpg:,.2f}")
        print(f"  the {len(bad)} units below burned {excess_gal:,.0f} gallons more than "
              f"they would have\n  at the better units' rate -- about "
              f"${excess_gal * price:,.0f} over {a.weeks} weeks, "
              f"${excess_gal * price / a.weeks * 52:,.0f} a year at this run rate.")


if __name__ == "__main__":
    main()
