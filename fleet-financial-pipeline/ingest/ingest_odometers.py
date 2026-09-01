"""
The odometer sheet -- five years of weekly readings, one row per unit.

Wide format: Unit | Driver | <date> | DIFFERENCE | <date> | DIFFERENCE | ...
266 weekly columns from Aug 2021 to Aug 2026, 300 unit rows. This is the
longest-running operational record in the corpus and the only mileage source
that spans the whole period, so utilization and cost-per-mile rest on it.

THE TRAP THAT MAKES NAIVE MILEAGE WRONG: one bad reading corrupts TWO weeks,
not one. If a unit reads 400,000 this week, 40,000 next (a transposition, a
replaced cluster, or a swapped unit), then next week's delta is -360,000 AND
the following week's is +360,000. Dropping only the negative leaves the
enormous positive behind and inflates mileage by exactly the amount you were
trying to remove. Both neighbours of a bad reading have to go.

So each delta is classified rather than filtered:
    ok          a plausible week
    negative    the odometer went backwards -- cluster swap, unit reassignment,
                or a mistyped reading
    implausible above MAX_WEEKLY_MILES; no truck runs that in a week
    tainted     arithmetically fine, but computed from a reading that one of
                the two rules above condemned

Only `ok` deltas are summed. The rest are kept and counted, because the number
of them is itself a data-quality measure -- and a unit with many is a unit
whose recorded mileage cannot be trusted for cost-per-mile.

The sheet also stores its own DIFFERENCE column, and it is FORWARD-LOOKING: the
column following a date holds the miles run from THAT date to the next one, not
the miles arriving at it. Reading it against the closing reading instead of the
opening one makes it disagree with the computed delta on 59% of weeks, which
looks like a data-quality disaster and is really just an off-by-one. Anchored
correctly the two agree, and the column is used only as a check.

Usage:
    python ingest/ingest_odometers.py Odometers.xlsx --outdir out/
"""
import argparse
import csv
import re
import warnings
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")

MAX_WEEKLY_MILES = 6000       # a hard ceiling: 70 mph x 11 h x 7 d is ~5,400
# Year width is inconsistent across the sheet: "08.02.21" for 2021-22 and
# 2025-26, but "01.02.2023" through the 2023-24 stretch, plus two typos with a
# three-digit year ("10.10.222"). A rule demanding exactly two digits silently
# drops 106 columns -- close to two years -- and the sheet then looks like it
# has a two-year hole in its records when it does not.
DATE_HDR = re.compile(r"^(\d{1,2})[.](\d{1,2})[.](\d{2,4})$")


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    return " ".join(str(v).split()) if v is not None else ""


def hdr_date(v):
    m = DATE_HDR.match(txt(v))
    if not m:
        return None
    mo, d, ys = int(m.group(1)), int(m.group(2)), m.group(3)
    # "222" is a slip for "22"; a 4-digit year is already absolute
    y = int(ys) if len(ys) == 4 else 2000 + int(ys[:2])
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--outdir", default="data/processed")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    out_rows = []

    for tab in wb.sheetnames:
        ws = wb[tab]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hi = next((i for i, r in enumerate(rows)
                   if any(hdr_date(c) for c in r[:12])), None)
        if hi is None:
            continue
        head = rows[hi]
        cols = [(j, hdr_date(head[j])) for j in range(len(head)) if hdr_date(head[j])]
        sheet_diff = {j: j + 1 for j, _ in cols}

        for r in rows[hi + 1:]:
            unit = txt(r[0] if r else "")
            if not unit:
                continue
            unit = unit[:-2] if unit.endswith(".0") else unit
            driver = txt(r[1]) if len(r) > 1 else ""

            series = [(d, num(r[j]) if j < len(r) else None,
                       num(r[sheet_diff[j]]) if sheet_diff[j] < len(r) else None)
                      for j, d in cols]
            # first pass: compute deltas between consecutive PRESENT readings
            pts = [(i, d, v) for i, (d, v, _) in enumerate(series) if v is not None]
            bad_idx = set()
            deltas = []
            for a, b in zip(pts, pts[1:]):
                (ia, da, va), (ib, db, vb) = a, b
                delta = vb - va
                if delta < 0:
                    flag = "negative"
                elif delta > MAX_WEEKLY_MILES * max(1, (db - da).days / 7):
                    flag = "implausible"
                else:
                    flag = "ok"
                if flag != "ok":
                    bad_idx.add(ia)
                    bad_idx.add(ib)
                deltas.append([ia, ib, da, db, va, vb, delta, flag])
            # second pass: a delta touching a condemned reading is tainted
            for d_ in deltas:
                if d_[7] == "ok" and (d_[0] in bad_idx or d_[1] in bad_idx):
                    d_[7] = "tainted"
            for ia, ib, da, db, va, vb, delta, flag in deltas:
                out_rows.append({
                    "tab": tab, "unit": unit, "driver": driver,
                    "week_start": da.isoformat(), "week_end": db.isoformat(),
                    "odo_open": va, "odo_close": vb,
                    "miles": round(delta, 1), "flag": flag,
                    # forward-looking: the DIFFERENCE after the OPENING date
                    "sheet_difference": series[ia][2]})

    out = Path(args.outdir)
    out.mkdir(parents=True, exist_ok=True)
    with (out / "odometers.csv").open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["tab", "unit", "driver", "week_start",
                                           "week_end", "odo_open", "odo_close",
                                           "miles", "flag", "sheet_difference"])
        w.writeheader()
        w.writerows(out_rows)

    ok = [r for r in out_rows if r["flag"] == "ok"]
    print(f"{len(out_rows):,} unit-week deltas from {len({r['unit'] for r in out_rows})} units")
    for f in ("ok", "negative", "implausible", "tainted"):
        n = sum(1 for r in out_rows if r["flag"] == f)
        mi = sum(r["miles"] for r in out_rows if r["flag"] == f)
        print(f"  {f:<13}{n:>7,}{mi:>16,.0f} miles")
    print(f"\n  USABLE MILEAGE: {sum(r['miles'] for r in ok):,.0f} over {len(ok):,} clean unit-weeks")
    if ok:
        print(f"  period {min(r['week_start'] for r in ok)} .. {max(r['week_end'] for r in ok)}")
    dis = [r for r in ok if r["sheet_difference"] is not None
           and abs(r["sheet_difference"] - r["miles"]) > 1]
    print(f"  sheet's own DIFFERENCE column disagrees with the computed delta on "
          f"{len(dis):,} of {len(ok):,} clean weeks")
    print(f"  -> {out / 'odometers.csv'}")


if __name__ == "__main__":
    main()
