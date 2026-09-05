"""Catalog every source file, so no session has to re-read the chat to find one.

WHY THIS EXISTS. The corpus is 235 files under `data/raw` plus whatever the
current session has been handed, and nothing about a path says what a file is:
`88206141-Xtrack_LLC_download.xlsx` is the weekly P&L, `4d409f87-Iron_lease.xlsx`
is a lease register that sits in TWO directories, and 130 bank PDFs across seven
accounts share about sixty filenames. Finding the right one by opening files
costs more than reading this catalog.

WHAT IT RECORDS, and the two things it is careful about:

1. IDENTITY IS THE CONTENT HASH, NOT THE NAME. BofA names every statement
   `eStmt_<period-end>.pdf`, so a name is not an identity -- see the
   full-path rule in CLAUDE.md. Files are keyed by sha256, which also surfaces
   the same document filed under two paths (there are such pairs) instead of
   letting it be counted twice.

2. THE CLASSIFIER IS ALLOWED TO SAY IT DOES NOT KNOW. A file it cannot place
   lands as kind `unknown` and is listed in the catalog under that heading. A
   guess that looks like knowledge is worse than a gap that is visible.

Run:  python ingest/catalog.py            # rewrite data/CATALOG.json + docs/CATALOG.md
      python ingest/catalog.py --check    # exit 1 if the catalog is out of date

A full build takes a few minutes: it opens every P&L workbook to record how many
of its weeks actually pass the gross control and whether any block-header
spelling is unknown to the reader. That is the point -- a tab count is not a
week count -- but it is why this is not instant. `--check` is fast.
"""
import argparse
import hashlib
import json
import re
import collections
import subprocess
import sys

import openpyxl
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from ingest_weekly_pnl import week_key, labeled, blocks, WANT
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "analysis"))
from xtrack_diagnosis import unmapped_headers

ROOT = Path(__file__).resolve().parent.parent
# Where source files live. Uploads are session-scoped and vanish with the
# container; data/raw is the working copy that survives inside a session.
SEARCH_ROOTS = ["data/raw", "data/processed"]
UPLOAD_ROOT = Path.home() / ".claude" / "uploads"

# NOTE the lookarounds rather than \b: '_' is a WORD character, so \bxtrack\b
# never matches 'Xtrack_LLC_download.xlsx' and every such file came back with no
# entity at all. The entity is read from the FILENAME first and only then from
# the directory, because data/raw/xtrack/ also holds Zone's and AFG's
# settlements -- matching on the path alone labelled all three XTRACK.
W = r"(?<![a-z0-9])%s(?![a-z0-9])"
ENTITY_PATTERNS = [
    (r"iron[_ ]?lease", "IRON_LEASE"),
    (r"truck[_ ]?max|truckmax|" + W % "shop", "TRUCKMAX"),
    (W % "xtrack", "XTRACK"),
    (W % "zone", "ZONE"),
    (W % "afg", "AFG"),
    (W % "stl", "STL"),
    (r"runstar", "RUNSTAR"),
]

# Order matters: the first match wins, so the specific sits above the general.
KIND_PATTERNS = [
    (r"\.(zip|rar|7z)$", "archive",
     "Upload archive. The extracted files are catalogued separately; the archive "
     "is kept only as the re-upload unit after a container reclaim."),
    (r"data/raw/ifta/", "fuel_tax_return",
     "IFTA quarterly and state weight-distance returns. FILED documents, so an "
     "independent record of miles and gallons -- identify the form by its Step 2 "
     "division line, not by the word 'IFTA', which some returns do not extract."),
    (r"data/raw/insurance/", "insurance_policy",
     "Signed policies, proposals and UCR registrations. Units are scheduled by "
     "VIN and cannot yet be joined to fleet numbers."),
    (r"data/raw/ops/", "ops_export",
     "Load-level export from the dispatch system: one row per load entry, with "
     "pickup/delivery city, gross, miles, dispatcher and pay type. The only source "
     "in the corpus below the weekly level."),
    (r"iron/invoices/", "lease_invoice",
     "Iron Lease's weekly invoices TO an operating company -- the billing side of "
     "the intercompany lease, which the bank alone cannot distinguish from funding."),
    (r"data/raw/amex/", "card_export",
     "AmEx card export. No beginning/ending balance -- control: none. Cross-check "
     "card payments against the verified bank statements instead."),
    (r"estmt_\d{4}-\d{2}-\d{2}\.pdf$", "bank_statement",
     "BofA statement PDF. Identity is the FULL PATH -- seven accounts share these names."),
    (r"-statements-\d{6}-", "driver_settlement",
     "Weekly driver settlement statements (one workbook per week)."),
    (r"profit.*loss|_download\.xlsx$|profit_and_loss", "pnl_weekly",
     "Weekly P&L workbook, ONE TAB PER WEEK; the week exists only in the tab name."),
    (r"efs_fuel_report", "fuel_efs",
     "EFS/WEX fuel rail. Carried all diesel through March 2026."),
    (r"relay", "fuel_relay",
     "Relay Payments fuel rail. Took over in April 2026. A rail itemises what "
     "later hits the bank as one draft -- reconcile, never sum."),
    (r"bestpass", "toll_bestpass", "Bestpass toll rail."),
    (r"truck_and_trailer_expenses|truck_max_charges", "shop_charges",
     "Shop / truck-and-trailer charge register."),
    (r"invoice_list", "factoring_invoices",
     "Triumph factoring invoice list. NO amount-received column -- a 'Short Paid' "
     "status carries the FULL invoice value, so it cannot measure what was collected."),
    (r"payroll", "payroll",
     "Payroll register. ADP is ONE rail for at least five companies, so an ADP "
     "total cannot be split by entity without this register."),
    (r"iron[_ ]?lease", "lease_register",
     "Iron Lease register. 'rented'/'sold' in Start date describe the DRIVER's "
     "standing, not Iron Lease's."),
    (r"oo_and_lo_drivers", "driver_roster", "Owner-operator and lease-operator roster."),
    (r"stmt (max truck|shop)|shop investments", "shop_charges",
     "Shop / truck-and-trailer charge register."),
    (r"bank_of_america.*\.csv$", "bank_feed_csv",
     "QuickBooks/BofA bank-feed CSV. NO beginning/ending balance, so it cannot be "
     "self-verified -- control: none. An export of exactly 300 rows hit the UI page "
     "cap and is a truncated view of the account, not the account."),
    (r"report__fb", "recruiting_cost",
     "Driver recruiting cost per candidate -- MVR, PSP, clearinghouse, drug test, "
     "travel -- with recruiter and hired/not-hired status. The acquisition side of "
     "the driver-turnover cost that never appears in the weekly P&L."),
    (r"^data/processed/", "processed_output",
     "Derived output of this pipeline, not a source. Regenerate it rather than "
     "reading it as evidence."),
    (r"\.csv$", "processed_csv", "Derived output of this pipeline, not a source."),
]


def sha256(path, chunk=1 << 20):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for b in iter(lambda: fh.read(chunk), b""):
            h.update(b)
    return h.hexdigest()


def classify(rel):
    s = str(rel).lower()
    kind, note = "unknown", ""
    for pat, k, n in KIND_PATTERNS:
        if re.search(pat, s):
            kind, note = k, n
            break
    name = s.rsplit("/", 1)[-1]
    entity = None
    for scope in (name, s):                       # filename wins over directory
        for pat, e in ENTITY_PATTERNS:
            if re.search(pat, scope):
                entity = e
                break
        if entity:
            break
    return kind, entity, note


PERIOD_PATTERNS = [
    (r"estmt_(\d{4}-\d{2}-\d{2})", lambda m: m.group(1)),
    (r"-statements-(\d{4})(\d{2})-", lambda m: f"wk{m.group(2)} of 20{m.group(1)[2:]}"),
    (r"\b(20\d{2})\b", lambda m: m.group(1)),
]


def period(rel):
    s = str(rel).lower()
    for pat, fn in PERIOD_PATTERNS:
        m = re.search(pat, s)
        if m:
            return fn(m)
    return None


def sheet_names(path, limit=6):
    """Tab names, which for the P&L workbooks ARE the time axis."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return len(names), names[:limit], week_span(names)
    except Exception:
        return None, [], None


def pnl_parse_status(path, tabs):
    """How much of a P&L workbook this pipeline can actually USE.

    Two different limits, and a tab count shows neither:

      weeks_passing_control  the panel's 'Total gross' equals the sum of that
                             week's unit blocks. Where it does not, the panel
                             figure was hand-set.
      unmapped_headers       block-header spellings ALIAS does not know. Those
                             columns read as ZERO -- silently -- so a workbook
                             with any is not safe to take costs from. The long
                             exports reach back to 2023 and use column names
                             ('Actual price') the later sheets dropped.

    This is why the XTRACK workbook with 145 tabs is not the one in use: its
    2026 weeks are identical to the 27-tab export, but its earlier years carry
    unmapped columns. Recorded here so no session re-derives it by parsing 145
    tabs.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        return {"weeks_passing_control": 0, "parse_note": f"{type(exc).__name__}: {exc}"}
    passing, failing, unmapped = [], 0, collections.Counter()
    for t in wb.sheetnames:
        wk = week_key(t)
        if not wk:
            continue
        ws = wb[t]
        unmapped += unmapped_headers(ws)
        panel = labeled(ws, WANT).get("gross")
        units = sum(g for g, _ in blocks(ws))
        if isinstance(panel, (int, float)) and panel and abs(units - panel) <= 1.0:
            passing.append(wk)
        else:
            failing += 1
    wb.close()
    out = {"weeks_passing_control": len(passing)}
    if failing:
        out["weeks_failing_control"] = failing
    if passing:
        out["passing_range"] = f"{min(passing)} .. {max(passing)}"
    if unmapped:
        out["unmapped_headers"] = dict(unmapped.most_common(8))
    return out


def week_span(names):
    """For a workbook whose tabs are weeks, the range those weeks cover."""
    wks = sorted(w for w in (week_key(n) for n in names) if w)
    return f"{wks[0]} .. {wks[-1]}" if wks else None


def consumers():
    """Which module reads which file, read off the code rather than maintained by hand."""
    out = {}
    try:
        hits = subprocess.run(
            ["grep", "-rn", "-o", r"data/raw/[A-Za-z0-9_./-]*", "analysis", "ingest", "tests",
             "--include=*.py"],
            cwd=ROOT, capture_output=True, text=True, check=False).stdout
    except OSError:
        return out
    for line in hits.splitlines():
        src, _, path = line.partition(":")[0], None, line.rsplit(":", 1)[-1]
        out.setdefault(path, set()).add(src)
    return {k: sorted(v) for k, v in out.items()}


def scan(include_uploads=True):
    files, roots = {}, [(ROOT / r, r) for r in SEARCH_ROOTS]  # sha256 -> [record, ...]
    if include_uploads and UPLOAD_ROOT.exists():
        for d in sorted(UPLOAD_ROOT.iterdir()):
            if d.is_dir():
                roots.append((d, f"~uploads/{d.name}"))
    used = consumers()
    for base, label in roots:
        if not base.exists():
            continue
        for p in sorted(base.rglob("*")):
            if "__MACOSX" in p.parts:          # macOS zip resource forks, not data
                continue
            if not p.is_file() or p.name in (".gitkeep", ".DS_Store"):
                continue
            rel = p.relative_to(base)
            shown = f"{label}/{rel}"
            digest = sha256(p)
            if digest in files:
                # Same bytes under another path -- the upload copy of a file
                # already in data/raw. Re-parsing it would double the cost of a
                # build for nothing; only the path is new.
                files[digest].append({"path": shown, "bytes": p.stat().st_size})
                continue
            kind, entity, note = classify(shown)
            rec = {"path": shown, "bytes": p.stat().st_size,
                   "modified": datetime.fromtimestamp(p.stat().st_mtime, timezone.utc)
                                       .strftime("%Y-%m-%d"),
                   "kind": kind, "entity": entity, "period": period(shown),
                   "note": note, "used_by": used.get(shown, [])}
            if p.suffix.lower() == ".csv":
                try:
                    with open(p, newline="") as fh:
                        rec["columns"] = fh.readline().strip().split(",")[:20]
                        rec["rows"] = sum(1 for _ in fh)
                except OSError:
                    pass
            if p.suffix.lower() in (".xlsx", ".xlsm"):
                n, names, span = sheet_names(p)
                if n is not None:
                    rec["sheets"] = n
                    rec["sheet_names"] = names
                    if span:
                        rec["period"] = span
                    if rec["kind"] == "pnl_weekly":
                        # The long exports carry 139-145 tabs; this is the slow
                        # part of a catalog build (a few minutes), so say so.
                        print(f"  parsing {rec['path']} ({n} tabs)...",
                              file=sys.stderr, flush=True)
                        rec.update(pnl_parse_status(p, n))
            files[digest] = [rec]
    return files


def build(files):
    entries = []
    for digest, copies in files.items():
        e = dict(copies[0])
        e["sha256"] = digest[:16]
        e["copies"] = [c["path"] for c in copies[1:]]
        entries.append(e)
    entries.sort(key=lambda e: (e["kind"], e["entity"] or "~", e["path"]))
    return {"generated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "distinct_files": len(entries),
            "total_paths": sum(1 + len(e["copies"]) for e in entries),
            "files": entries}


HEADINGS = {
    "pnl_weekly": "Weekly P&L workbooks",
    "bank_statement": "Bank statement PDFs",
    "driver_settlement": "Driver settlement statements",
    "fuel_efs": "Fuel — EFS/WEX rail",
    "fuel_relay": "Fuel — Relay rail",
    "toll_bestpass": "Tolls — Bestpass",
    "shop_charges": "Shop / truck & trailer charges",
    "factoring_invoices": "Factoring invoice lists",
    "payroll": "Payroll",
    "lease_register": "Iron Lease register",
    "driver_roster": "Driver rosters",
    "fuel_tax_return": "IFTA and state fuel-tax returns",
    "insurance_policy": "Insurance policies and registrations",
    "ops_export": "Dispatch system export (load level)",
    "lease_invoice": "Iron Lease invoices to the operating companies",
    "card_export": "AmEx card exports",
    "recruiting_cost": "Driver recruiting costs",
    "bank_feed_csv": "Bank-feed CSV exports",
    "processed_output": "Derived (pipeline output, not a source)",
    "processed_csv": "Derived (pipeline output, not a source)",
    "archive": "Upload archives (re-upload units)",
    "unknown": "Unclassified — open one of these before quoting it",
}


def markdown(cat):
    L = [
        "# Source file catalog",
        "",
        "Generated by `ingest/catalog.py` -- **do not edit by hand**, rerun it.",
        "",
        f"{cat['distinct_files']} distinct files by content hash "
        f"({cat['total_paths']} paths; the difference is the same document filed twice). "
        f"Catalog built {cat['generated']}.",
        "",
        "`~uploads/` is the current session's upload directory: it lives in the",
        "container and does NOT survive the container being reclaimed. `data/raw`",
        "is gitignored for the same reason it is sensitive -- see **Durability**",
        "at the foot of this file.",
        "",
    ]
    by_kind = {}
    for e in cat["files"]:
        by_kind.setdefault(e["kind"], []).append(e)
    order = list(HEADINGS) + [k for k in sorted(by_kind) if k not in HEADINGS]
    for kind in order:
        rows = by_kind.get(kind)
        if not rows:
            continue
        L += [f"## {HEADINGS.get(kind, kind)}", ""]
        note = next((r["note"] for r in rows if r["note"]), "")
        if note:
            L += [f"> {note}", ""]
        if kind == "ops_export":
            L += ["| file | rows | columns |", "|---|--:|---|"]
            for r in rows:
                L += [f"| `{r['path'].split('/')[-1]}` | {r.get('rows', '')} | "
                      f"{', '.join('`%s`' % c for c in r.get('columns', []))} |"]
            L += [""]
            continue
        if kind in ("bank_statement", "driver_settlement", "lease_invoice",
                    "card_export", "bank_feed_csv", "fuel_tax_return",
                    "insurance_policy"):
            # Hundreds of near-identical files: summarise, never list.
            grouped = {}
            for r in rows:
                grouped.setdefault(r["path"].rsplit("/", 1)[0], []).append(r)
            L += ["| directory | files | entity | period range |", "|---|--:|---|---|"]
            for d, rs in sorted(grouped.items()):
                per = sorted(p for p in (r["period"] for r in rs) if p)
                span = f"{per[0]} .. {per[-1]}" if per else "—"
                ent = sorted({r["entity"] for r in rs if r["entity"]}) or ["—"]
                L += [f"| `{d}` | {len(rs)} | {'/'.join(ent)} | {span} |"]
            L += [""]
            continue
        pnl = kind == "pnl_weekly"
        head = ("| file | entity | period | tabs | weeks passing control | read by |"
                if pnl else "| file | entity | period | tabs | read by |")
        L += [head, "|---|---|---|--:|--:|---|" if pnl else "|---|---|---|--:|---|"]
        for r in rows:
            tabs = str(r.get("sheets", "")) or ""
            used = ", ".join(f"`{u}`" for u in r["used_by"]) or "—"
            dup = f"<br>also at `{r['copies'][0]}`" if r["copies"] else ""
            cells = [f"`{r['path']}`{dup}", r["entity"] or "—", r["period"] or "—", tabs]
            if pnl:
                rd = str(r.get("weeks_passing_control", ""))
                if r.get("passing_range") and r.get("weeks_failing_control"):
                    rd += f"<br><small>{r['passing_range']}</small>"
                if r.get("unmapped_headers"):
                    rd += ("<br>**unmapped: "
                           + ", ".join(f"`{k}`" for k in r["unmapped_headers"]) + "**")
                cells.append(rd)
            cells.append(used)
            L += ["| " + " | ".join(cells) + " |"]
        L += [""]
    stranded = [e for e in cat["files"]
                if all(p.startswith("~uploads/") for p in [e["path"]] + e["copies"])]
    loose = [e for e in stranded if e["kind"] != "archive"]
    archives = [e for e in stranded if e["kind"] == "archive"]
    if loose:
        L += ["## Uploaded but not filed into `data/raw`", "",
              "These reached the session and were never unpacked or moved into the",
              "working tree, so no analysis reads them. Either file them or say why not.",
              f"({len(archives)} upload archives are also unfiled by design -- their",
              "contents are already in the tree and they are kept as the re-upload unit.)",
              "", "| file | kind | entity |", "|---|---|---|"]
        for e in sorted(loose, key=lambda e: (e["kind"], e["path"])):
            L += [f"| `{e['path'].split('/')[-1]}` | {e['kind']} | {e['entity'] or '—'} |"]
        L += [""]

    L += [
        "## Durability",
        "",
        "This container is ephemeral and has already been reclaimed once mid-analysis,",
        "taking `data/raw` and `data/processed` with it. `data/raw` is gitignored",
        "(bank statements and payroll), so **the catalog is committed and the files",
        "are not**. After a reclaim, this file is the list of what has to come back:",
        "re-upload against it and rerun `python ingest/catalog.py --check` to prove",
        "the corpus is whole before quoting a number from it.",
        "",
    ]
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--check", action="store_true",
                    help="exit 1 if the committed catalog no longer matches the files")
    ap.add_argument("--no-uploads", action="store_true")
    a = ap.parse_args()

    cat = build(scan(include_uploads=not a.no_uploads))
    js, md = ROOT / "data/CATALOG.json", ROOT / "docs/CATALOG.md"
    if a.check:
        try:
            old = json.loads(js.read_text())
        except OSError:
            print("no catalog committed yet; run without --check", file=sys.stderr)
            return 1
        def paths(entry):
            return [entry["path"]] + entry.get("copies", [])

        def keep(entry):
            # With --no-uploads the session upload directory is not scanned, so
            # a file that only ever existed there is not "missing" -- it is out
            # of scope for this run. A file that also has a data/raw copy is.
            return not a.no_uploads or any(not p.startswith("~uploads/") for p in paths(entry))

        was = {e["sha256"]: e["path"] for e in old["files"] if keep(e)}
        now = {e["sha256"]: e["path"] for e in cat["files"]}
        gone = {k: v for k, v in was.items() if k not in now}
        new = {k: v for k, v in now.items() if k not in was}
        for k, v in sorted(gone.items()):
            print(f"MISSING  {v}")
        for k, v in sorted(new.items()):
            print(f"NEW      {v}")
        print(f"{len(now)} present, {len(gone)} missing, {len(new)} new")
        return 1 if (gone or new) else 0

    js.write_text(json.dumps(cat, indent=1) + "\n")
    md.parent.mkdir(exist_ok=True)
    md.write_text(markdown(cat))
    print(f"{cat['distinct_files']} distinct files ({cat['total_paths']} paths) "
          f"-> {js.relative_to(ROOT)}, {md.relative_to(ROOT)}")
    unknown = [e["path"] for e in cat["files"] if e["kind"] == "unknown"]
    if unknown:
        print(f"unclassified ({len(unknown)}):")
        for p in unknown:
            print(f"  {p}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
