"""
Read the consolidated weekly P&L workbooks (ZONE, XTRACK, AFG) tab by tab.

ONE TAB PER WEEK, and the week exists ONLY in the tab name -- no cell carries a
date. So the tab name is the time axis and must never be dropped or renamed.

TWO TRAPS THIS MODULE EXISTS TO AVOID

1. The overhead labels differ between workbooks. ZONE writes 'Salary of US
   office' in the panel header; XTRACK writes 'US salary' there and puts
   'Salary of US office' ~18 rows further down in a per-mile block. Matching on
   the label alone lands on the wrong cell in XTRACK and returns a POSITIVE
   number for a cost. So the overhead block is anchored on the ROW carrying
   'Tas_team salaries' -- present in both -- and read from the row beneath it.

2. Panel positions move between workbooks and between weeks. 'Total gross' sits
   at row 2 in ZONE and row 20 in XTRACK. Nothing may be read by fixed
   coordinates; every field is found by its label.

THE TWO MILEAGE FIGURES, settled by the operator 2026-09-03. They are NOT two
measures of the same thing -- they differ in population AND in what counts as a
mile, which is why they sit about 2x apart at XTRACK:

  'Total Odometer mileage'  COMPANY DRIVERS ONLY, and ALL miles those trucks
                            ran -- loaded and empty. Owner-operator and
                            lease-to-own trucks are absent.
  'Total mileage'           ALL trucks, company plus OO plus LO, but LOADED
                            miles only.

So neither is "the" mileage, and subtracting one from the other means nothing.
Each metric has to take the denominator that matches what it measures:

  revenue per mile, cost per mile   -> 'Total mileage' (loaded, all trucks),
                                       because revenue is earned on loaded miles
                                       and every truck earns it
  fuel economy and fuel cost/mile   -> odometer miles, because fuel is burned on
                                       every mile including empty ones, and the
                                       company only buys fuel for its own drivers
  deadhead                          -> a company truck's odometer against its own
                                       loaded miles. Measured over the last four
                                       weeks: ZONE 6.6%, XTRACK 7.6%, AFG 10.4%.

CONTROL: the panel's 'Total gross' must equal the sum of the week's unit blocks
(each unit is a block of load rows closed by a row reading 'Total' in column B).
Run check_weekly_pnl() on every new export before quoting anything from it.
"""
import openpyxl
import re
import json
import sys
DATE=re.compile(r'(\d{1,2})\.(\d{1,2})\.(\d{2})\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{2})')
def week_key(t):
    m=DATE.search(t.replace(' ',''))
    return f"20{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}" if m else None
def num(v): return float(v) if isinstance(v,(int,float)) else None

OH_ALIAS={'salary of us office':'us_office','us salary':'us_office',"owner's":'owners',
          'tas_team salaries':'tashkent','other charges':'other','net profit':'net_profit'}

def overhead_row(ws):
    """Anchor on the ROW that carries 'Tas_team salaries'; values sit directly below.
    Labels differ between workbooks ('Salary of US office' vs 'US salary'), and
    'Salary of US office' ALSO appears further down in a per-mile block -- matching
    the label alone lands on the wrong cell and returns a positive number."""
    for r in ws.iter_rows(min_row=1,max_row=20,min_col=15,max_col=24):
        labs={c.column:c.value.strip().lower() for c in r
              if isinstance(c.value,str) and c.value.strip()}
        if 'tas_team salaries' in labs.values():
            out={}
            for col,lab in labs.items():
                k=OH_ALIAS.get(lab)
                if k: out[k]=num(ws.cell(row=r[0].row+1,column=col).value)
            return out
    return {}

def labeled(ws,wanted,maxr=40):
    """Label in col P-ish, value in the next column. Only within the panel block."""
    out={}
    for r in ws.iter_rows(min_row=1,max_row=maxr,min_col=16,max_col=21):
        for c in r:
            if isinstance(c.value,str):
                k=wanted.get(c.value.strip().lower())
                if k and k not in out:
                    v=num(ws.cell(row=c.row,column=c.column+1).value)
                    if v is not None: out[k]=v
    return out

WANT={'total gross':'gross','total odometer mileage':'odo_miles','total mileage':'all_miles',
      'total driver pay':'driver_pay','total fuel':'fuel','total truck rent':'truck_rent',
      'total toll and scale':'tolls','total toll':'tolls','other expenses total':'other_exp',
      'salaries uzbekistan':'uz_salaries','salaries of oo':'oo_salaries',
      'expenses of oo':'oo_expenses','insurance':'insurance','oo trucks':'oo_trucks',
      'cd trucks':'cd_trucks','average rpm (company)':'rpm_company',
      'average rpm (oo)':'rpm_oo','fuel discount for oo':'oo_fuel_discount'}

def blocks(ws):
    t=[]
    for r in range(1,ws.max_row+1):
        if str(ws.cell(row=r,column=2).value).strip()=='Total':
            t.append((num(ws.cell(row=r,column=3).value) or 0.0,
                      num(ws.cell(row=r,column=4).value) or 0.0))
    return t

def truck_counts(ws):
    """'OO trucks'/'CD trucks' labels sit left of their count in the far-right block."""
    out={}
    for r in ws.iter_rows(min_row=25,max_row=40,min_col=19,max_col=22):
        for c in r:
            if isinstance(c.value,str) and c.value.strip().lower() in ('oo trucks','cd trucks'):
                out[c.value.strip().lower().replace(' ','_')]=num(
                    ws.cell(row=c.row,column=c.column+1).value)
    return out

WORKBOOKS={"ZONE":"data/raw/pnl/4954206d-Zone_LLC_download.xlsx",
           "XTRACK":"data/raw/pnl/88206141-Xtrack_LLC_download.xlsx",
           "AFG":"data/raw/pnl/423b1a54-AFG__download.xlsx"}

def read_workbook(path):
    """Every week in one workbook, keyed by week. Import-safe: nothing at module
    level opens a file, so importing this module costs nothing."""
    wb=openpyxl.load_workbook(path,data_only=True); out={}
    for t in wb.sheetnames:
        k=week_key(t)
        if not k: continue
        ws=wb[t]; d={'tab':t}
        d.update(overhead_row(ws)); d.update(labeled(ws,WANT)); d.update(truck_counts(ws))
        b=blocks(ws); d['unit_blocks']=len(b)
        d['unit_gross']=sum(x[0] for x in b); d['unit_miles']=sum(x[1] for x in b)
        out[k]=d
    return out


def check_weekly_pnl(weeks, tol=1.0):
    """Panel gross vs the sum of unit blocks. Returns (passes, failures)."""
    ok, bad = [], []
    for wk, d in sorted(weeks.items()):
        panel, units = d.get("gross"), d.get("unit_gross")
        if not isinstance(panel, (int, float)) or panel == 0:
            bad.append((wk, panel, units, None))
            continue
        diff = units - panel
        (ok if abs(diff) <= tol else bad).append((wk, panel, units, diff))
    return ok, bad


if __name__=="__main__":
    out={lab:read_workbook(f) for lab,f in WORKBOOKS.items()}
    for lab,w in out.items(): print(f"{lab}: {len(w)} weeks",file=sys.stderr)
    json.dump(out,open(sys.argv[1],'w'),indent=1)
