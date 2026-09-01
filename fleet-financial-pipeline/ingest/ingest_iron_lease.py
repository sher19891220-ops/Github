"""
Iron Lease workbook -- the leasing entity's own billing and truck register.

Two things live in every "Overall <entity>" tab, side by side but NOT row
aligned. Reading them as one table produces nonsense, because they are simply
two independent lists that happen to share a sheet:

  cols A-E   the TRUCK REGISTER: truck number, driver, start date, pickup
             mileage. One row per truck. "rented" in the date/mileage columns
             marks a unit Iron Lease rents in rather than owns.
  cols G-J   the WEEKLY INVOICE LEDGER: date period, invoice amount, paid
             amount, date paid. One row per WEEK, not per truck.

PAYMENT LAGS INVOICE BY ONE ROW. On the AFG ledger, the week invoiced $3,294.60
shows $4,167.16 paid; the next week invoices $157.39 and shows $3,294.60 paid --
last week's invoice. They settle a week in arrears, consistently. So a row-wise
invoice-vs-paid comparison is meaningless and the period totals are what matter.

The 57 weekly tabs are a different shape again: per entity, per truck, an
opening and closing ODOMETER reading, the mileage difference, a mileage charge,
weeks worked, truck rent and an EFS/maintenance column. That makes this the
only source in the corpus with a real odometer reading per truck per week, and
it is what the rent-vs-mileage billing is actually computed from.

Usage:
    python ingest/ingest_iron_lease.py Iron_lease.xlsx --outdir out/
"""
import argparse
import csv
import re
import warnings
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")

PERIOD = re.compile(r"(\d{1,2})[.](\d{1,2})[.](\d{2})\s*-\s*(\d{1,2})[.](\d{1,2})[.](\d{2})")
ENTITY_HDR = re.compile(r"^(ZONE|XTRACK|AFG|IRON\s*LEASE|TRUCK\s*MAX)\s*:?\s*$", re.I)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if not s or s.lower() in ("rented", "accident", "n/a", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    return str(v).strip() if v is not None else ""


def week_of(s):
    m = PERIOD.search(str(s).replace(" ", ""))
    if not m:
        return "", ""
    a, b, c, d, e, f = (int(x) for x in m.groups())
    try:
        return date(2000 + c, a, b).isoformat(), date(2000 + f, d, e).isoformat()
    except ValueError:
        return "", ""


def parse_overall(ws, entity):
    """Returns (trucks, invoice_ledger) -- two independent lists."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = next((i for i, r in enumerate(rows)
                if any(txt(c) == "Truck number" for c in r[:6])), None)
    if hdr is None:
        return [], []
    h = {txt(c).lower(): j for j, c in enumerate(rows[hdr]) if txt(c)}
    tcol = h.get("truck number")
    dcol = h.get("driver")
    scol = h.get("start date")
    mcol = h.get("pu mileage")
    pcol = h.get("date period")
    icol = h.get("invoice amount")
    paycol = h.get("paid amount")
    dpcol = h.get("date paid")

    trucks, ledger = [], []
    for r in rows[hdr + 1:]:
        g = lambda j: (r[j] if j is not None and j < len(r) else None)
        tn = txt(g(tcol))
        if tn:
            trucks.append({"entity": entity, "truck": tn, "driver": txt(g(dcol)),
                           "start_date": txt(g(scol)),
                           "pu_mileage": num(g(mcol)),
                           # "rented" in these columns means Iron Lease does not
                           # own the unit -- it is renting it in and re-letting it
                           "rented_in": "yes" if txt(g(scol)).lower() == "rented" else "no"})
        per = txt(g(pcol))
        inv, paid = num(g(icol)), num(g(paycol))
        if per and (inv is not None or paid is not None):
            w0, w1 = week_of(per)
            ledger.append({"entity": entity, "period": per, "week_start": w0,
                           "week_end": w1, "invoiced": inv or 0.0,
                           "paid": paid or 0.0, "date_paid": txt(g(dpcol))})
    return trucks, ledger


def parse_week_tab(ws, tab):
    """Per-truck odometer + rent rows, grouped under an entity heading.

    Tabs before the entity split carry NO heading at all -- Zone was the only
    operating company then, which the billing ledger corroborates (the Xtrack
    and AFG ledgers both begin 2026-05-04, and the pre-split "Overall" ledger
    restates ZONE's weeks identically). Those rows are labelled ZONE with
    entity_source=inferred_presplit, so an inference is never mistaken for
    something read off the sheet."""
    out, entity = [], ""
    w0, w1 = week_of(tab)
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        # The entity heading sits in column B, not column A.
        head = next((txt(c) for c in r[:3] if txt(c) and ENTITY_HDR.match(txt(c))), "")
        if head:
            entity = head.rstrip(":").upper().replace(" ", "")
            continue
        first = txt(r[0])
        if first == "#" or not first:
            continue
        truck = txt(r[1]) if len(r) > 1 else ""
        if not truck or not truck[0].isdigit():
            continue
        vals = [num(x) for x in r[2:10]] + [None] * 8
        # The first tab is an INCEPTION sheet: it carries each truck's
        # cumulative odometer and to-date charges, not one week's activity.
        # Its rows have no opening reading, and taken as weekly they contribute
        # 1.6M of 3.4M "miles" and $1.64M of $1.83M of "mileage charge" -- a
        # truck does not run 316,620 miles in a week. A row without an opening
        # reading cannot yield a delta, so it is recorded with no miles rather
        # than with a fabricated one.
        if vals[0] is None and vals[1] is not None:
            vals[2] = vals[3] = None
        out.append({"entity": entity or "ZONE",
                    "entity_source": "heading" if entity else "inferred_presplit",
                    "tab": tab, "week_start": w0, "week_end": w1,
                    "truck": truck, "odo_open": vals[0], "odo_close": vals[1],
                    "miles": vals[2], "mileage_charge": vals[3],
                    "worked_weeks": vals[4], "truck_rent": vals[5],
                    "efs_maintenance": vals[6]})
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--outdir", default="data/processed")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    trucks, ledger, weeks = [], [], []
    for tab in wb.sheetnames:
        ws = wb[tab]
        if tab.lower().startswith("overall"):
            ent = tab[7:].strip().upper() or "ALL"
            t, l = parse_overall(ws, ent)
            trucks += t
            ledger += l
        elif PERIOD.search(tab.replace(" ", "")):
            weeks += parse_week_tab(ws, tab)

    out = Path(args.outdir)
    out.mkdir(parents=True, exist_ok=True)
    for name, rows, cols in (
        ("iron_lease_trucks", trucks,
         ["entity", "truck", "driver", "start_date", "pu_mileage", "rented_in"]),
        ("iron_lease_ledger", ledger,
         ["entity", "period", "week_start", "week_end", "invoiced", "paid", "date_paid"]),
        ("iron_lease_weekly", weeks,
         ["entity", "entity_source", "tab", "week_start", "week_end", "truck",
          "odo_open", "odo_close", "miles", "mileage_charge", "worked_weeks",
          "truck_rent", "efs_maintenance"]),
    ):
        with (out / f"{name}.csv").open("w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        print(f"  {name:<22}{len(rows):>7,} rows")

    print(f"\n{len([t for t in wb.sheetnames if PERIOD.search(t.replace(' ', ''))])} weekly tabs parsed")


if __name__ == "__main__":
    main()
