"""The unit-to-VIN map, and therefore which company runs which insured truck.

This closes the gap every insurance question was blocked on. Policies schedule
units by VIN; the P&Ls, the driver roster, the maintenance ledger and the Iron
Lease register all use fleet numbers, and the two share no visible relationship
-- only 5 of the 68 VINs on the master auto-liability policy contain a fleet
number anywhere in them. The group unit workbook carries both columns, so it is
the join.

WHAT THE UNIT WORKBOOK ACTUALLY IS. Not a fleet list: an ASSIGNMENT HISTORY.
One row per driver who has had the truck, with pick-up and drop-off mileage and
dates. ZONE's sheet has 1,137 rows carrying a VIN and 294 distinct VINs, one
truck appearing up to 15 times. Counting rows counts assignments and multiplies
the fleet fourfold.

WHICH COMPANY OWNS A TRUCK IS NOT IN THIS WORKBOOK EITHER. 35 of the 362
distinct VINs appear on more than one company's sheet, because a truck that
moved between authorities stays in both histories. The weekly P&Ls settle it:
whichever company's P&L last carried that unit is the company running it. That
also dates the answer, which matters -- 249 of the 362 VINs appear in no P&L at
all and are history, not fleet.

VALUES ARE FREE TEXT. The Value column holds numbers, '110K$', '$60,000/OO' and
bare 'OO'. 'OO' marks an owner-operator truck, which carries no company physical
damage. money() reads the lot and flags the OO ones rather than dropping them.
"""
import argparse
import collections
import datetime
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "ingest"))
sys.path.insert(0, str(ROOT / "analysis"))

WORKBOOK = ROOT / "data/raw/be9f8984-group_units_list_and_driver_list.xlsx"
UNIT_SHEETS = [(" Zone Unit List", "ZONE"), ("Xtrack_Unit ", "XTRACK"),
               ("AFG Unit list ", "AFG")]
VIN_LEN = 17
# A VIN is 17 chars from a restricted alphabet -- I, O and Q are never used. The
# column also carries section headers ('NEW WALMART TRUCKS'), which are the right
# length to slip through a bare length test.
VIN_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")


def money(v):
    """(value, is_owner_operator). '110K$', '$60,000/OO' and bare 'OO' all appear."""
    if isinstance(v, (int, float)):
        return float(v), False
    s = str(v or "").strip()
    if not s or s.lower() == "none":
        return None, False
    oo = "oo" in s.lower()
    m = re.search(r"([\d,]+(?:\.\d+)?)\s*[kK]?", s.replace("$", ""))
    if not m:
        return None, oo
    x = float(m.group(1).replace(",", ""))
    if re.search(r"\d\s*[kK]", s) or 10 < x < 1000:      # '110K$' and '110$K'
        x *= 1000
    return x, oo


def read_units():
    """Every row of every unit sheet, keyed by VIN."""
    import openpyxl
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    out = collections.defaultdict(lambda: {"units": set(), "lists": set(),
                                           "value": None, "owner_operator": False,
                                           "make": None, "model": None, "year": None})
    rows_seen = 0
    for sheet, company in UNIT_SHEETS:
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        hdr = [str(c).strip() if c else "" for c in rows[0]]
        ix = {h: i for i, h in enumerate(hdr)}
        ucol = ix.get("Unit Number", ix.get("Unit", 1))
        vcol = ix.get("Vin Number", 6)
        for r in rows[1:]:
            vin = str(r[vcol] or "").strip().upper()
            if not VIN_RE.match(vin):
                continue
            rows_seen += 1
            u = str(r[ucol] or "").strip()
            u = u[:-2] if u.endswith(".0") else u
            rec = out[vin]
            rec["lists"].add(company)
            if u:
                rec["units"].add(u)
            val, oo = money(r[5] if len(r) > 5 else None)
            if val and not rec["value"]:
                rec["value"] = val
            rec["owner_operator"] = rec["owner_operator"] or oo
            for k, i in (("make", 2), ("model", 3), ("year", 4)):
                if rec[k] is None and len(r) > i:
                    rec[k] = r[i]
    return dict(out), rows_seen


def last_seen_in_pnl():
    """unit -> (week, company, block kind) of the most recent P&L that carried it."""
    import openpyxl
    from ingest_weekly_pnl import week_key, WORKBOOKS
    from xtrack_diagnosis import read_blocks
    last = {}
    for company, path in WORKBOOKS.items():
        wb = openpyxl.load_workbook(ROOT / path, data_only=True)
        for tab in wb.sheetnames:
            wk = week_key(tab)
            if not wk:
                continue
            for b in read_blocks(wb[tab]):
                u = b["unit"]
                if u not in last or wk > last[u][0]:
                    last[u] = (wk, company, b["kind"])
    return last


def registry():
    units, rows_seen = read_units()
    last = last_seen_in_pnl()
    out = []
    for vin, rec in units.items():
        hits = sorted(((last[u][0], last[u][1], last[u][2], u)
                       for u in rec["units"] if u in last), reverse=True)
        out.append({"vin": vin, "units": sorted(rec["units"]),
                    "unit": hits[0][3] if hits else (sorted(rec["units"])[0]
                                                     if rec["units"] else None),
                    "on_lists": sorted(rec["lists"]),
                    "company": hits[0][1] if hits else None,
                    "last_week": hits[0][0] if hits else None,
                    "kind": hits[0][2] if hits else None,
                    "value": rec["value"], "owner_operator": rec["owner_operator"],
                    "make": rec["make"], "model": rec["model"], "year": rec["year"]})
    return out, rows_seen


def controls(reg, rows_seen):
    fails = []
    if len(reg) >= rows_seen:
        fails.append(("the workbook is being read as a fleet list, not an "
                      "assignment history", f"{len(reg)} VINs from {rows_seen} rows"))
    multi = [r for r in reg if len(r["on_lists"]) > 1]
    if not multi:
        fails.append(("no VIN appears on two company sheets -- the histories may "
                      "have been split, so company attribution needs rechecking", 0))
    bad = [r["vin"] for r in reg if not VIN_RE.match(r["vin"])]
    if bad:
        fails.append(("values in the VIN column that are not VINs", bad[:5]))
    if any(r["value"] and r["value"] < 1000 for r in reg):
        fails.append(("stated values below $1,000 -- a 'K' suffix was not expanded",
                      [r["unit"] for r in reg if r["value"] and r["value"] < 1000][:5]))
    return fails


def active(reg, since="2026-07-06"):
    return [r for r in reg if r["last_week"] and r["last_week"] >= since]


def insured_value_by_company(reg, since="2026-07-06", exclude=()):
    """The physical-damage allocation basis: stated value of the company-owned
    trucks each company is actually running. Owner-operator units carry their
    own physical damage and are excluded."""
    g = collections.defaultdict(lambda: {"units": 0, "value": 0.0})
    for r in active(reg, since):
        if r["owner_operator"] or not r["value"] or r["company"] in exclude:
            continue
        g[r["company"]]["units"] += 1
        g[r["company"]]["value"] += r["value"]
    total = sum(x["value"] for x in g.values())
    for x in g.values():
        x["share"] = x["value"] / total if total else 0.0
    return dict(g), total


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--json", default="data/processed/fleet_registry.json")
    ap.add_argument("--since", default="2026-07-06")
    a = ap.parse_args()
    reg, rows_seen = registry()
    print(f"{rows_seen:,} assignment rows -> {len(reg)} distinct VINs")
    for what, v in controls(reg, rows_seen):
        print(f"  CONTROL: {what}: {v}")
    act = active(reg, a.since)
    print(f"  {len(act)} seen in a P&L since {a.since}; "
          f"{sum(1 for r in reg if not r['last_week'])} appear in no P&L at all")
    print(f"  {sum(1 for r in reg if len(r['on_lists']) > 1)} VINs sit on more than "
          f"one company's sheet (trucks that moved authority)")

    print("\n== ACTIVE COMPANY-OWNED FLEET, THE PHYSICAL-DAMAGE BASIS ==")
    g, total = insured_value_by_company(reg, a.since)
    for co, x in sorted(g.items(), key=lambda kv: -kv[1]["value"]):
        print(f"  {str(co):<9}{x['units']:>4} units  ${x['value']:>12,.0f}  {100 * x['share']:>6.1f}%")
    print(f"  {'TOTAL':<9}{sum(x['units'] for x in g.values()):>4} units  ${total:>12,.0f}")

    g2, t2 = insured_value_by_company(reg, a.since, exclude=("AFG",))
    print("\n== EXCLUDING AFG (the group physical-damage policy does not cover it) ==")
    for co, x in sorted(g2.items(), key=lambda kv: -kv[1]["value"]):
        print(f"  {str(co):<9}{x['units']:>4} units  ${x['value']:>12,.0f}  {100 * x['share']:>6.1f}%")
    print(f"  {'TOTAL':<9}{sum(x['units'] for x in g2.values()):>4} units  ${t2:>12,.0f}")

    out = ROOT / a.json
    out.parent.mkdir(parents=True, exist_ok=True)
    json.dump(reg, out.open("w"), indent=1, default=str)
    print(f"\n-> {a.json}")


if __name__ == "__main__":
    main()
