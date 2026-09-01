"""
"Cases" -- the breakdown and road-service log, one tab per month.

Every unplanned event that stopped a truck: the unit, the driver, what broke,
which shop fixed it, whether it was handled In Shop, on a Road Call or needed
Towing, how it was paid, and who the cost was charged to. It is the only source
that says WHY a truck was down, and the Service column is the severity scale --
a road call and a tow are not maintenance, they are a failure that already cost
a load.

DO NOT USE THE "Statistics" TAB. Its July and August EFS figures read
$1,601,244,272 and $2,295,179,848, and its grand total $3,896,690,877 -- invoice
numbers pasted into an amount column. The monthly detail tabs are sound and are
what this reads; the summary built on top of them is not.

Cost is split across four columns naming who bore it: Zone, STL, Driver, Dealer.
As in the truck-and-trailer ledger, that split is the recovery question -- what
the company absorbed versus what it passed on.

Amounts arrive with trailing commas ("969.90,", "1914.30,") from hand entry, so
parsing has to strip them rather than fail.

Usage:
    python ingest/ingest_cases.py Cases2026.xlsx --outdir out/
"""
import argparse
import csv
import re
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]
PAYERS = ["zone", "stl", "driver", "dealer"]

# No single roadside event costs this much. Above it, the cell is an invoice
# number, not a price.
IMPLAUSIBLE_EVENT = 50_000


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    return " ".join(str(v).split()) if v is not None else ""


def isodate(v):
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = txt(v)
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:19], f).date().isoformat()
        except ValueError:
            pass
    return ""


def parse_month(ws, month):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = next((i for i, r in enumerate(rows)
                if any(txt(c).lower().startswith("issue") for c in r[:6])), None)
    if hdr is None:
        return []
    h = {}
    for j, c in enumerate(rows[hdr]):
        k = txt(c).lower().replace("\n", " ").strip()
        if k and k not in h:
            h[k] = j
    out = []
    for r in rows[hdr + 1:]:
        g = lambda k: (r[h[k]] if k in h and h[k] < len(r) else None)
        unit = txt(g("unit number")) or txt(r[0] if r else "")
        issue = txt(g("issue"))
        if not issue and not unit:
            continue
        rec = {"month": month, "unit": unit.rstrip(".0") if unit.endswith(".0") else unit,
               "driver": txt(g("driver's info or dr")) or txt(g("driver's info")),
               "issue": issue[:160], "date": isodate(g("date")),
               "status": txt(g("status")), "shop": txt(g("shop's name")),
               "service": txt(g("service")), "payment": txt(g("payment")),
               "note": txt(g("note"))[:80], "invoice": txt(g("invoice number")),
               "bill": txt(g("bill"))}
        total = 0.0
        for p in PAYERS:
            v = num(g(p)) or 0.0
            rec[f"cost_{p}"] = v
            total += v
        rec["cost_total"] = round(total, 2)
        # Seven rows carry an INVOICE NUMBER in a cost column -- a tire and
        # brake-drum job priced at $2,295,073,865. They are what corrupts the
        # workbook's own Statistics tab. Flagged, never silently summed: an
        # implausible cost is a data-entry error to fix, not a number to keep.
        rec["cost_suspect"] = "yes" if total > IMPLAUSIBLE_EVENT else "no"
        if rec["cost_suspect"] == "yes":
            rec["cost_total"] = 0.0
            for p2 in PAYERS:
                rec[f"cost_{p2}"] = 0.0
        if rec["issue"] or total:
            out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--outdir", default="data/processed")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    rows = []
    for m in MONTHS:
        if m in wb.sheetnames:
            r = parse_month(wb[m], m)
            rows += r
            if r:
                print(f"  {m:<12}{len(r):>6} events  "
                      f"${sum(x['cost_total'] for x in r):>12,.2f}")
    out = Path(args.outdir)
    out.mkdir(parents=True, exist_ok=True)
    cols = (["month", "date", "unit", "driver", "issue", "service", "status", "shop",
             "payment", "note", "invoice", "bill"]
            + [f"cost_{p}" for p in PAYERS] + ["cost_total", "cost_suspect"])
    with (out / "cases.csv").open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    sus = [r for r in rows if r["cost_suspect"] == "yes"]
    print(f"\n{len(rows):,} breakdown events -> {out / 'cases.csv'}")
    print(f"  ${sum(r['cost_total'] for r in rows):,.2f} of costed work")
    if sus:
        print(f"  {len(sus)} rows zeroed as data-entry errors (invoice number in a "
              f"cost cell); units: {', '.join(sorted({r['unit'] for r in sus}))}")
    print("  (the workbook's own Statistics tab is not used: its July/August "
          "figures are invoice numbers, not dollars)")


if __name__ == "__main__":
    main()
