"""
Weekly per-unit P&L workbooks (the "<Entity> Profit and Loss Weekly" sheets).

One Google Sheet per operating company, ONE TAB PER WEEK, named for the week it
covers ("08.17.26-08.23.26"). Each tab holds two independent things side by
side, and both matter:

  left       repeated per-unit blocks: a "Unit# | Driver | Gross | Mileage |
             Driver Salary | ... " header, one row per LOAD, then a "Total" row
             that is the unit's week. The unit number is in column A of the
             first row of the block; the rows under it are that unit's loads.
  right      a weekly summary panel of label/value pairs -- "Total gross",
             "Total driver pay", "Salaries of OO", "C drivers" / "OO drivers".
             This panel is where the owner-operator split lives, and it exists
             nowhere else in the corpus.

NEITHER IS AT A FIXED COLUMN. Across three years the sheet was rebuilt several
times: 10 of ZONE's 139 tabs drop "Other Charges" and "Total" from the unit
table, so "Per mile" moves from index 13 to 11; the summary panel sits at
column P in recent tabs and column N in 2024 ones; and the same field is headed
"Insur/Admin/Trl" in most tabs and "Pys/Cargo/Admin" in others. Reading by
POSITION silently returns the wrong field -- truck rental read out of the toll
column, and a summary panel that simply is not found, which is what "panel
0.00" against real unit rows means. So every column here is resolved by
matching the header TEXT in each block, and every panel metric is found by
anchoring on its label and taking the first number to its right.

Two things to know before trusting anything this produces:

1. THE WEEK IS ONLY IN THE TAB NAME. The cell contents carry no date at all, so
   a text export of this workbook loses the time axis entirely. The tab name is
   the sole source of the period -- which is why this reads .xlsx and not the
   text rendering. Tab names are hand-typed and carry stray spaces and mixed
   separators; parse_week() is deliberately forgiving about that and refuses
   rather than guesses when it cannot read one.

2. THESE ARE ASSERTIONS, NOT CASH. This is the P&L the business has been
   deciding on -- hand-maintained, and the thing the bank data exists to test.
   Values are kept in the sheet's own orientation (costs positive) and are NOT
   sign-flipped here; normalisation happens at comparison time so that a
   mismatch is visible rather than absorbed. Formula cells that evaluated to
   #DIV/0! or #REF! are recorded as null, never as zero: a zero would silently
   average into a per-mile figure.

Usage:
    python ingest/ingest_gsheet_pnl.py AFG_PnL.xlsx --entity AFG --outdir out/
"""
import argparse
import csv
import re
import warnings
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")

UNIT_COLS = ["unit", "driver", "gross", "mileage", "driver_salary",
             "insur_admin_trl", "def_fuel_fee", "truck_rental", "toll_scale",
             "additional_charges", "subtotal", "other_charges", "total",
             "per_mile", "fuel_avr"]

# Header text -> canonical field. Same column, different wording by vintage.
HEADER_ALIASES = {
    "unit#": "unit", "unit": "unit", "driver": "driver", "gross": "gross",
    "mileage": "mileage", "driver salary": "driver_salary",
    "insur/admin/trl": "insur_admin_trl", "pys/cargo/admin": "insur_admin_trl",
    "insur/admin/trailer": "insur_admin_trl",
    "def/fuel/fee": "def_fuel_fee", "truck rental": "truck_rental",
    "toll / scale": "toll_scale", "toll/scale": "toll_scale",
    "additional charges": "additional_charges", "subtotal": "subtotal",
    "other charges": "other_charges", "total": "total",
    "per mile": "per_mile", "fuel avr": "fuel_avr",
}

# Labels in the right-hand weekly panel worth keeping. Matched case-folded and
# anchored, so a panel that moved columns is still found.
PANEL_LABELS = {
    "total gross", "total mileage", "total odometer mileage", "total driver pay",
    "total fuel", "total truck rent", "total toll and scale", "total fuel oo",
    "other expenses total", "salaries uzbekistan", "fuel discount for oo",
    "salary of us office", "salaries of oo", "expenses of oo", "insurance",
    "margin % (profit)", "average gross", "average rpm", "avr gross for running trucks",
    "maintance exp per truck", "maintance exp per mile", "toll average per mile",
    "fuel average per mile", "gallon", "worked units", "oo trucks", "cd trucks",
    "fuel discount oo", "total fuel oo", "loss-making trucks",
}

# Bands whose values sit on the row BELOW the labels, spread across columns.
BAND_HEADS = {"c drivers", "us salary", "salary"}

# "08.17.26-08.23.26", " 04.27.26- 05.03.26", "07.20.26 - 07.26.26"
WEEK = re.compile(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})\s*[-–]\s*"
                  r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})")

ERR = ("#DIV/0!", "#REF!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!")


def parse_week(tab):
    t = tab.replace(" ", "")
    # Four Xtrack tabs are typed "03.24.25.25-03.30.25" -- the year keyed twice.
    # Anchored to a FOUR-group date, because "01.26.26" is a legitimate
    # Jan 26 2026 whose day and year happen to match, and a looser rule
    # collapses it to "01.26" and loses three real weeks.
    t = re.sub(r"\b(\d{1,2}\.\d{1,2}\.\d{2})\.\d{2}(?=-)", r"\1", t)
    m = WEEK.search(t)
    if not m:
        return None, None
    a, b, c, d, e, f = (int(x) for x in m.groups())
    y1 = c + 2000 if c < 100 else c
    y2 = f + 2000 if f < 100 else f
    try:
        return date(y1, a, b), date(y2, d, e)
    except ValueError:
        return None, None


def num(v):
    """Formula errors become None, never 0.0 -- a zero would average in."""
    if v is None or (isinstance(v, str) and (not v.strip() or v.strip() in ERR)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if s in ERR or not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    try:
        x = float(s.strip("()"))
    except ValueError:
        return None
    return -x if neg else x


def txt(v):
    return str(v).strip() if v is not None else ""


def _colmap(r):
    """Build {canonical field -> column index} from an actual header row."""
    m = {}
    for j, v in enumerate(r):
        key = txt(v).lower()
        if key in HEADER_ALIASES and HEADER_ALIASES[key] not in m:
            m[HEADER_ALIASES[key]] = j
    return m


def _panel(rows, i, r, first_free):
    """Label-anchored metric extraction from the right-hand panel.

    Takes the first numeric cell within three columns to the right of a known
    label. Three, not one: the panel is merged and padded differently in each
    vintage, so the value is not always in the adjacent cell."""
    out = []
    for j in range(first_free, len(r)):
        lab = txt(r[j])
        if not lab or num(lab) is not None:
            continue
        low = lab.lower()
        if low in PANEL_LABELS:
            for k in range(j + 1, min(j + 4, len(r))):
                v = num(r[k])
                if v is not None:
                    out.append((lab, v))
                    break
        elif low in BAND_HEADS and i + 1 < len(rows):
            nxt = rows[i + 1]
            for k in range(j, min(j + 6, len(r))):
                sub = txt(r[k])
                if not sub or num(sub) is not None:
                    continue
                v = num(nxt[k]) if k < len(nxt) else None
                if v is not None:
                    out.append((sub, v))
    return out


def parse_sheet(ws, entity, tab):
    """Returns (unit_week_rows, weekly_metric_rows)."""
    w0, w1 = parse_week(tab)
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    units, metrics = [], []
    cur_unit = cur_driver = None
    in_block = False
    cmap, first_free = {}, 15

    for i, r in enumerate(rows):
        def c(i_):
            return r[i_] if i_ < len(r) else None

        head = txt(c(0))
        # --- per-unit blocks -------------------------------------------------
        if head == "Unit#":
            in_block, cur_unit, cur_driver = True, None, None
            cmap = _colmap(r)
            # everything right of the unit table is panel territory
            first_free = max(cmap.values()) + 1 if cmap else 15
            continue
        for lab, v in _panel(rows, i, r, first_free):
            metrics.append({"entity": entity, "tab": tab,
                            "week_start": w0.isoformat() if w0 else "",
                            "week_end": w1.isoformat() if w1 else "",
                            "metric": lab, "value": v})
        if in_block:
            if cur_unit is None and head and head != "Unit#":
                cur_unit = txt(c(0)).rstrip(".0") if txt(c(0)).endswith(".0") else head
                cur_driver = txt(c(1))
            if txt(c(1)) == "Total":
                rec = {"entity": entity, "tab": tab,
                       "week_start": w0.isoformat() if w0 else "",
                       "week_end": w1.isoformat() if w1 else "",
                       "unit": cur_unit or "", "driver": cur_driver or ""}
                for name in UNIT_COLS[2:]:
                    j = cmap.get(name)
                    rec[name] = num(c(j)) if j is not None else None
                # Each tab ships ~82 pre-built empty blocks as scaffolding for
                # a fleet larger than the one that ran. Keep a block only if it
                # names a unit or driver, or moved money. An idle truck with
                # gross 0 and a negative total is NOT empty -- that is rent and
                # insurance accruing on a truck that did not run, which is
                # exactly the kind of cost this analysis exists to find.
                vals = [v for k, v in rec.items()
                        if k in UNIT_COLS[2:] and v not in (None, 0.0)]
                if rec["unit"] or rec["driver"] or vals:
                    units.append(rec)
                in_block = False

    return units, metrics


def parse_split_rows(ws, entity, tab):
    """The 'C drivers / OO drivers / Discount / Total' and
    'US salary / Owner's / Tas_team salaries / Other charges / Net profit'
    bands are headers with their values on the FOLLOWING row."""
    w0, w1 = parse_week(tab)
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    out = []
    for i, r in enumerate(rows[:-1]):
        labels = [txt(r[j]) if j < len(r) else "" for j in range(15, 21)]
        if labels[0] in ("C drivers", "US salary"):
            nxt = rows[i + 1]
            for k, lab in enumerate(labels):
                if not lab:
                    continue
                v = num(nxt[15 + k]) if 15 + k < len(nxt) else None
                if v is not None:
                    out.append({"entity": entity, "tab": tab,
                                "week_start": w0.isoformat() if w0 else "",
                                "week_end": w1.isoformat() if w1 else "",
                                "metric": lab, "value": v})
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--entity", required=True)
    ap.add_argument("--outdir", default="data/processed")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    all_units, all_metrics, undated = [], [], []
    for tab in wb.sheetnames:
        ws = wb[tab]
        u, m = parse_sheet(ws, args.entity, tab)
        all_units += u
        all_metrics += m
        if parse_week(tab) == (None, None):
            undated.append(tab)

    out = Path(args.outdir)
    out.mkdir(parents=True, exist_ok=True)
    up = out / f"pnl_unit_week_{args.entity}.csv"
    mp = out / f"pnl_metrics_{args.entity}.csv"
    with up.open("w", newline="") as fh:
        cols = ["entity", "tab", "week_start", "week_end", "unit", "driver"] + UNIT_COLS[2:]
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(all_units)
    with mp.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["entity", "tab", "week_start",
                                           "week_end", "metric", "value"])
        w.writeheader()
        w.writerows(all_metrics)

    weeks = sorted({r["week_start"] for r in all_units if r["week_start"]})
    print(f"{args.entity}: {len(wb.sheetnames)} tabs -> "
          f"{len(all_units):,} unit-weeks, {len(all_metrics):,} weekly metrics")
    if weeks:
        print(f"  period: {weeks[0]} .. {max(r['week_end'] for r in all_units if r['week_end'])}")
    if undated:
        print(f"  {len(undated)} tab(s) with no readable week -- NOT dated, "
              f"reported rather than guessed: {', '.join(undated[:6])}")
    print(f"  -> {up}\n  -> {mp}")


if __name__ == "__main__":
    main()
