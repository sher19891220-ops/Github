"""Tests for the weekly P&L reader, especially the two label traps."""
import sys
from pathlib import Path
import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent / "ingest"))
from ingest_weekly_pnl import week_key, overhead_row, check_weekly_pnl, blocks


def test_week_comes_from_the_tab_name_including_padded_variants():
    assert week_key("08.17.26-08.23.26") == "2026-08-17"
    assert week_key(" 04.27.26- 05.03.26") == "2026-04-27"
    assert week_key("3.2.26-3.8.26") == "2026-03-02"
    assert week_key("Sheet1") is None


def _panel(labels, values):
    wb = openpyxl.Workbook(); ws = wb.active
    for i, l in enumerate(labels):
        ws.cell(row=12, column=16 + i, value=l)
    for i, v in enumerate(values):
        ws.cell(row=13, column=16 + i, value=v)
    return ws


def test_zone_label_spelling():
    ws = _panel(["Salary of US office", "Owner's", "Tas_team salaries",
                 "Other charges", "Net profit"],
                [-4500, -2000, -12445, -23368.35, 23043.20])
    d = overhead_row(ws)
    assert d["us_office"] == -4500 and d["tashkent"] == -12445


def test_xtrack_label_spelling_maps_to_the_same_field():
    ws = _panel(["US salary", "Owner's", "Tas_team salaries",
                 "Other charges", "Net profit"],
                [-4500, -500, -14365, -42091.33, -14844.03])
    d = overhead_row(ws)
    assert d["us_office"] == -4500          # 'US salary', not 'Salary of US office'
    assert d["tashkent"] == -14365


def test_the_decoy_label_further_down_is_never_read():
    """XTRACK repeats 'Salary of US office' ~18 rows below the panel with a
    POSITIVE per-mile figure. Anchoring on the label instead of the row picks
    that up and turns a cost into income."""
    ws = _panel(["US salary", "Tas_team salaries"], [-4500, -14365])
    ws.cell(row=30, column=16, value="Salary of US office")
    ws.cell(row=30, column=17, value=4500)          # the decoy, positive
    d = overhead_row(ws)
    assert d["us_office"] == -4500


def test_overhead_absent_returns_empty_rather_than_guessing():
    wb = openpyxl.Workbook()
    assert overhead_row(wb.active) == {}


def test_unit_blocks_are_the_rows_closed_by_Total_in_column_B():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=5, column=2, value="Total"); ws.cell(row=5, column=3, value=1000)
    ws.cell(row=5, column=4, value=500)
    ws.cell(row=9, column=2, value="Total"); ws.cell(row=9, column=3, value=2000)
    ws.cell(row=9, column=4, value=800)
    ws.cell(row=7, column=3, value=99999)           # a load row, must not count
    b = blocks(ws)
    assert len(b) == 2
    assert sum(x[0] for x in b) == 3000
    assert sum(x[1] for x in b) == 1300


def test_control_passes_when_panel_matches_units():
    ok, bad = check_weekly_pnl({"2026-01-05": {"gross": 1000.0, "unit_gross": 1000.0}})
    assert len(ok) == 1 and not bad


def test_control_fails_when_panel_exceeds_units():
    ok, bad = check_weekly_pnl({"2026-01-05": {"gross": 1000.0, "unit_gross": 990.0}})
    assert not ok and bad[0][3] == pytest.approx(-10.0)


def test_control_flags_a_week_with_no_panel_rather_than_passing_it():
    ok, bad = check_weekly_pnl({"2026-01-05": {"gross": None, "unit_gross": 500.0}})
    assert not ok and bad[0][3] is None
